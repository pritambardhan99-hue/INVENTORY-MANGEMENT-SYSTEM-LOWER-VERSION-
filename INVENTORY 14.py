import os
import re
import sqlite3
import datetime as dt
from dataclasses import dataclass
from typing import Optional, Tuple, List, Any
from matplotlib.ticker import FuncFormatter


import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from PIL import Image, ImageTk, ImageEnhance
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, Table, TableStyle, SimpleDocTemplate, Spacer

import pandas as pd
from matplotlib import pyplot as plt
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

APP_TITLE = "Inventory Management System"

THEME = {
    "primary": "#1ABC9C",  # teal
    "dark": "#2C3E50",  # navy
    "accent": "#E67E22",  # orange
    "bg": "#F7F9FA",
    "text": "#2C3E50",
    "danger": "#E74C3C",
    "success": "#27AE60",
    "warning": "#F1C40F"
}
FONT_LG = ("Segoe UI", 14)
FONT_XL = ("Segoe UI", 18, "bold")
FONT_MD = ("Segoe UI", 12)

DB_PATH = "inventory14.db"


# ---------- Helpers ----------

def db() -> sqlite3.Connection:
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    return con


def init_db():
    con = db()
    cur = con.cursor()

    # Users
    cur.execute("""
                CREATE TABLE IF NOT EXISTS users
                (
                    user_id
                    INTEGER
                    PRIMARY
                    KEY
                    AUTOINCREMENT,
                    username
                    TEXT
                    UNIQUE
                    NOT
                    NULL,
                    password
                    TEXT
                    NOT
                    NULL,
                    role
                    TEXT
                    CHECK (
                    role
                    IN
                (
                    'Admin',
                    'Employee'
                )) NOT NULL,
                    is_online INTEGER DEFAULT 0,
                    last_login TEXT
                    );
                """)

    # Employees
    cur.execute("""
                CREATE TABLE IF NOT EXISTS employees
                (
                    emp_id
                    TEXT
                    PRIMARY
                    KEY,
                    name
                    TEXT
                    NOT
                    NULL,
                    phone
                    TEXT
                    UNIQUE
                    NOT
                    NULL,
                    email
                    TEXT
                    UNIQUE
                    NOT
                    NULL,
                    role
                    TEXT
                    NOT
                    NULL,
                    join_date
                    TEXT
                    NOT
                    NULL
                );
                """)

    # Suppliers
    cur.execute("""
                CREATE TABLE IF NOT EXISTS suppliers
                (
                    supplier_id
                    TEXT
                    PRIMARY
                    KEY,
                    name
                    TEXT
                    NOT
                    NULL,
                    company
                    TEXT
                    NOT
                    NULL,
                    phone
                    TEXT
                    UNIQUE
                    NOT
                    NULL,
                    email
                    TEXT
                    UNIQUE
                    NOT
                    NULL,
                    address
                    TEXT
                );
                """)

    # Customers
    cur.execute("""
                CREATE TABLE IF NOT EXISTS customers
                (
                    customer_id
                    TEXT
                    PRIMARY
                    KEY,
                    name
                    TEXT
                    NOT
                    NULL,
                    phone
                    TEXT
                    UNIQUE
                    NOT
                    NULL,
                    email
                    TEXT
                    UNIQUE
                    NOT
                    NULL,
                    address
                    TEXT
                );
                """)

    # Products
    cur.execute("""
                CREATE TABLE IF NOT EXISTS products
                (
                    product_id
                    TEXT
                    PRIMARY
                    KEY, -- SKU like 001
                    name
                    TEXT
                    NOT
                    NULL,
                    category
                    TEXT
                    NOT
                    NULL,
                    supplier_id
                    TEXT
                    NOT
                    NULL,
                    quantity
                    INTEGER
                    NOT
                    NULL
                    DEFAULT
                    0,
                    gst            REAL NOT NULL DEFAULT 18,
                    unit_price
                    REAL
                    NOT
                    NULL
                    DEFAULT
                    0.0, -- NEW FIELD
                    mrp
                    REAL
                    NOT
                    NULL,
                    reorder_level
                    INTEGER
                    NOT
                    NULL
                    DEFAULT
                    0,
                    FOREIGN
                    KEY
                (
                    supplier_id
                ) REFERENCES suppliers
                (
                    supplier_id
                )
                    );
                """)

    # Sales
    cur.execute("""
                CREATE TABLE IF NOT EXISTS sales
                (
                    sale_id
                    INTEGER
                    PRIMARY
                    KEY
                    AUTOINCREMENT,
                    product_id
                    TEXT
                    NOT
                    NULL,
                    product_name
                    TEXT
                    NOT
                    NULL,
                    category
                    TEXT,
                    quantity
                    INTEGER
                    NOT
                    NULL,
                    mrp
                    REAL
                    NOT
                    NULL,
                    total_price
                    REAL
                    NOT
                    NULL, -- qty * mrp (before discount)
                    discount_type
                    TEXT, -- "Flat" or "Percent"
                    discount_value
                    REAL, -- discount value (₹ or %)
                    effective_total
                    REAL
                    NOT
                    NULL, -- final price after discount
                    date
                    TEXT
                    NOT
                    NULL, -- sale date (YYYY-MM-DD)
                    sold_by
                    TEXT, -- employee username
                    customer_name
                    TEXT,
                    customer_phone
                    TEXT
                );


                """)
    # ♻️ Returns
    cur.execute("""
                   CREATE TABLE IF NOT EXISTS returns
                   (
                       return_id
                       INTEGER
                       PRIMARY
                       KEY
                       AUTOINCREMENT,
                       sale_id
                       INTEGER,
                       product_id
                       TEXT,
                       quantity
                       INTEGER,
                       refund_amount
                       REAL,
                       date
                       TEXT,
                       reason
                       TEXT,
                       FOREIGN
                       KEY
                   (
                       sale_id
                   ) REFERENCES sales
                   (
                       sale_id
                   ),
                       FOREIGN KEY
                   (
                       product_id
                   ) REFERENCES products
                   (
                       product_id
                   )
                       )
                   """)
    cur.execute("""
                   CREATE TABLE IF NOT EXISTS returns
                   (
                       return_id
                       INTEGER
                       PRIMARY
                       KEY
                       AUTOINCREMENT,
                       sale_id
                       INTEGER,
                       product_id
                       TEXT,
                       quantity
                       INTEGER,
                       refund_amount
                       REAL,
                       date
                       TEXT,
                       reason
                       TEXT,
                       FOREIGN
                       KEY
                   (
                       sale_id
                   ) REFERENCES sales
                   (
                       sale_id
                   ),
                       FOREIGN KEY
                   (
                       product_id
                   ) REFERENCES products
                   (
                       product_id
                   )
                       )
                   """)

    # Seed admin if missing
    cur.execute("SELECT 1 FROM users WHERE username=?", ("admin",))
    if cur.fetchone() is None:
        cur.execute("INSERT INTO users(username,password,role,is_online,last_login) VALUES(?,?,?,?,?)",
                    ("admin", "admin123", "Admin", 0, None))
    con.commit()
    con.close()


def padded_id(prefix_table: str, id_col: str, width: int = 3) -> str:
    """
    Generate next numeric string ID (e.g., 001, 002).
    prefix_table: table name
    id_col: id column
    """
    con = db()
    cur = con.cursor()
    cur.execute(f"SELECT {id_col} FROM {prefix_table}")
    ids = []
    for row in cur.fetchall():
        try:
            ids.append(int(str(row[0]).lstrip("0") or "0"))
        except Exception:
            pass
    nxt = (max(ids) + 1) if ids else 1
    con.close()
    return str(nxt).zfill(width)


def validate_email(email: str) -> bool:
    # allow only gmail.com or yahoo.com
    return re.fullmatch(r"^[A-Za-z0-9._%+-]+@(gmail\.com|yahoo\.com)$", email) is not None


def validate_phone(phone: str) -> bool:
    return re.fullmatch(r"^[6-9]\d{9}$", phone) is not None


def today_str() -> str:
    return dt.date.today().isoformat()


def now_str() -> str:
    return dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def employee_default_password(emp_name: str) -> str:
    token = re.sub(r"\s+", "", emp_name).lower()[:3]
    if len(token) < 3:
        token = (token + "xxx")[:3]
    return f"{token}123"


# ---------- PDF: Invoice ----------

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet


def generate_invoice_pdf(filename, company_name, company_address, invoice_no, invoice_date,
                         customer_name, customer_phone, items, discount_type, discount_value,
                         gst_percent, subtotal, grand_total):
    """
    Generate invoice PDF with full breakdown.
    items = list of (name, category, qty, mrp, line_total)
    """

    doc = SimpleDocTemplate(filename, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=20)
    story = []
    styles = getSampleStyleSheet()

    # Header
    story.append(Paragraph(f"<b>{company_name}</b>", styles["Title"]))
    story.append(Paragraph(company_address.replace("\n", "<br/>"), styles["Normal"]))
    story.append(Spacer(1, 12))

    # Invoice meta
    story.append(Paragraph(f"<b>Invoice No:</b> {invoice_no}", styles["Normal"]))
    story.append(Paragraph(f"<b>Date:</b> {invoice_date}", styles["Normal"]))
    story.append(Paragraph(f"<b>Customer:</b> {customer_name}", styles["Normal"]))
    story.append(Paragraph(f"<b>Phone:</b> {customer_phone}", styles["Normal"]))
    story.append(Spacer(1, 12))

    # Table header
    data = [["Product", "Category", "Qty", "MRP", "Line Total (₹)"]]
    for n, c, q, m, t in items:
        data.append([n, c, q, f"{m:.2f}", f"{t:.2f}"])

    table = Table(data, colWidths=[180, 100, 60, 80, 100])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightblue),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story.append(table)
    story.append(Spacer(1, 12))

    # Totals breakdown
    totals_data = []
    totals_data.append(["Subtotal", f"₹ {subtotal:.2f}"])
    if discount_type == "Flat":
        totals_data.append([f"Discount (Flat ₹{discount_value:.2f})", f"- ₹ {discount_value:.2f}"])
    else:
        totals_data.append([f"Discount ({discount_value:.2f}%)", f"- ₹ {(subtotal * discount_value / 100):.2f}"])
    after_disc = subtotal - (discount_value if discount_type == "Flat" else subtotal * discount_value / 100)
    gst_amt = after_disc * (gst_percent / 100)
    totals_data.append([f"GST ({gst_percent:.1f}%)", f"+ ₹ {gst_amt:.2f}"])
    totals_data.append(["", ""])
    totals_data.append(["Grand Total", f"₹ {grand_total:.2f}"])

    totals_table = Table(totals_data, colWidths=[300, 200])
    totals_table.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
        ("FONTNAME", (0, 0), (-1, -2), "Helvetica"),
        ("FONTNAME", (-1, -1), (-1, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (-1, -1), (-1, -1), colors.green),
        ("FONTSIZE", (-1, -1), (-1, -1), 14),
    ]))
    story.append(totals_table)

    doc.build(story)


# ---------- Excel / PDF generic exports ----------

def export_treeview_to_excel(tree: ttk.Treeview, suggested_name: str):
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=suggested_name,
                                             filetypes=[("Excel Workbook", "*.xlsx")])
    if not save_path:
        return
    cols = tree["columns"]
    data = []
    for child in tree.get_children():
        vals = tree.item(child, "values")
        data.append(list(vals))
    df = pd.DataFrame(data, columns=cols)
    try:
        with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Data")
        messagebox.showinfo("Export", f"Excel exported:\n{save_path}")
    except Exception as e:
        messagebox.showerror("Export Error", str(e))


def export_treeview_to_pdf(tree: ttk.Treeview, suggested_name: str, title: str):
    save_path = filedialog.asksaveasfilename(defaultextension=".pdf", initialfile=suggested_name,
                                             filetypes=[("PDF", "*.pdf")])
    if not save_path:
        return

    cols = tree["columns"]
    data = [list(cols)]
    for child in tree.get_children():
        vals = tree.item(child, "values")
        data.append(list(vals))

    doc = SimpleDocTemplate(save_path, pagesize=A4, rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    story = [Paragraph(f"<b>{title}</b>", styles["Title"]), Spacer(1, 8)]

    tbl = Table(data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.black),
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONT', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
    ]))
    story.append(tbl)
    doc.build(story)
    messagebox.showinfo("Export", f"PDF exported:\n{save_path}")


# ---------- Main App ----------

class InventoryApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1200x750")
        self.configure(bg=THEME["bg"])
        self.minsize(1100, 720)

        self.current_user = None  # (username, role)

        self.container = tk.Frame(self, bg=THEME["bg"])
        self.container.pack(fill="both", expand=True)

        self.login_frame = LoginFrame(self.container, self)
        self.login_frame.pack(fill="both", expand=True)

        self.protocol("WM_DELETE_WINDOW", self.on_close)

    def show_dashboard(self):
        self.login_frame.pack_forget()
        self.dashboard = Dashboard(self.container, self)
        self.dashboard.pack(fill="both", expand=True)

    def logout(self):
        # mark offline
        if self.current_user:
            con = db()
            cur = con.cursor()
            cur.execute("UPDATE users SET is_online=0 WHERE username=?", (self.current_user[0],))
            con.commit()
            con.close()

        for w in self.container.winfo_children():
            w.destroy()
        self.current_user = None
        self.login_frame = LoginFrame(self.container, self)
        self.login_frame.pack(fill="both", expand=True)

    def on_close(self):
        try:
            if self.current_user:
                con = db()
                cur = con.cursor()
                cur.execute("UPDATE users SET is_online=0 WHERE username=?", (self.current_user[0],))
                con.commit()
                con.close()
        except:
            pass
        self.destroy()



from PIL import Image, ImageTk   # <-- for loading logo images

class LoginFrame(tk.Frame):
    def __init__(self, parent, app: InventoryApp):
        super().__init__(parent, bg=THEME["bg"])
        self.app = app
        self.attempts = 0  # Track invalid login attempts

        # --- Main two-column layout ---
        main_frame = tk.Frame(self, bg=THEME["bg"])
        main_frame.place(relx=0.5, rely=0.5, anchor="center")

        # Left column → logo2.png
        try:
            logo2_img = Image.open("logo2.png")  # keep logo2.png in same folder
            logo2_img = logo2_img.resize((550, 650))  # adjust size if needed
            self.logo2 = ImageTk.PhotoImage(logo2_img)
            tk.Label(main_frame, image=self.logo2, bg=THEME["bg"]).grid(
                row=0, column=0, padx=(0, 40), pady=10, sticky="n"
            )
        except Exception as e:
            print("Logo2 not found:", e)

        # Right column → login form
        wrapper = tk.Frame(main_frame, bg=THEME["bg"])
        wrapper.grid(row=0, column=1, sticky="n")

        # --- Logo above title ---
        try:
            logo_img = Image.open("logo.png")
            logo_img = logo_img.resize((180, 180))
            self.logo = ImageTk.PhotoImage(logo_img)
            tk.Label(wrapper, image=self.logo, bg=THEME["bg"]).grid(row=0, column=0, columnspan=2, pady=(0, 10))
        except Exception as e:
            print("Logo not found:", e)

        # Title
        title = tk.Label(wrapper, text=APP_TITLE, font=FONT_XL, fg=THEME["dark"], bg=THEME["bg"])
        title.grid(row=1, column=0, columnspan=2, pady=(0, 16))

        # Username
        tk.Label(wrapper, text="Username", font=FONT_LG, bg=THEME["bg"]).grid(
            row=2, column=0, sticky="e", padx=8, pady=6
        )
        usr_holder = tk.Frame(wrapper, bg="#FFF59D", bd=2, relief="flat")
        usr_holder.grid(row=2, column=1, sticky="w", padx=8, pady=6)

        self.username_var = tk.StringVar()
        self.username_cmb = ttk.Combobox(
            usr_holder, textvariable=self.username_var, state="readonly", width=25
        )
        self.username_cmb.pack(padx=4, pady=2)

        # Password
        tk.Label(wrapper, text="Password", font=FONT_LG, bg=THEME["bg"]).grid(
            row=3, column=0, sticky="e", padx=8, pady=6
        )
        pwd_holder = tk.Frame(wrapper, bg="#C8E6C9", bd=2, relief="flat")
        pwd_holder.grid(row=3, column=1, sticky="w", padx=8, pady=6)
        self.password_var = tk.StringVar()
        self.password_entry = tk.Entry(
            pwd_holder, textvariable=self.password_var, show="•", font=FONT_LG, bd=0, width=25
        )
        self.password_entry.pack(padx=4, pady=2)

        # Clock
        self.clock_lbl = tk.Label(wrapper, text="", font=FONT_MD, fg="red", bg=THEME["bg"])
        self.clock_lbl.grid(row=4, column=0, columnspan=2, pady=8)
        self.update_clock()

        # Login Button
        btn = tk.Button(wrapper, text="Login", font=FONT_LG, bg=THEME["primary"], fg="white",
                        activebackground=THEME["accent"], height=2,
                        command=self.try_login, cursor="hand2")
        btn.grid(row=5, column=0, columnspan=2, sticky="ew", pady=8)



        # load usernames
        self.refresh_usernames()

    def update_clock(self):
        self.clock_lbl.config(text=dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        self.after(1000, self.update_clock)

    def refresh_usernames(self):
        con = db()
        cur = con.cursor()
        cur.execute("SELECT username FROM users ORDER BY username")
        users = [r[0] for r in cur.fetchall()]
        con.close()
        self.username_cmb["values"] = users
        if users:
            self.username_cmb.current(0)

    def try_login(self):
        username = self.username_var.get().strip()
        password = self.password_var.get().strip()
        if not username or not password:
            messagebox.showerror("Login", "Please select username and enter password.")
            return

        con = db()
        cur = con.cursor()
        cur.execute("SELECT username,password,role FROM users WHERE username=?", (username,))
        row = cur.fetchone()
        if not row or row["password"] != password:
            messagebox.showerror("Login", "Invalid credentials.")
            con.close()
            return

        # mark online + last_login
        cur.execute("UPDATE users SET is_online=1, last_login=? WHERE username=?", (now_str(), username))
        con.commit()
        con.close()

        self.app.current_user = (row["username"], row["role"])
        self.app.show_dashboard()

# ---------- Dashboard & Sections ----------
class Dashboard(tk.Frame):
    def __init__(self, parent, app: InventoryApp):
            super().__init__(parent, bg=THEME["bg"])

            self.app = app
            self.current_section_frame: Optional[tk.Frame] = None

            # Header
            header = tk.Frame(self, bg=THEME["dark"], height=60)
            header.pack(side="top", fill="x")

            tk.Label(header, text=APP_TITLE, font=FONT_XL, fg="white", bg=THEME["dark"]).pack(side="left", padx=16)
            self.dt_lbl = tk.Label(header, text=now_str(), font=FONT_MD, fg="white", bg=THEME["dark"])
            self.dt_lbl.pack(side="left", padx=16)
            self.update_header_clock()

            user_txt = f"{self.app.current_user[0]} ({self.app.current_user[1]})"
            tk.Label(header, text=user_txt, font=FONT_LG, fg="white", bg=THEME["dark"]).pack(side="right", padx=16)

            logout_btn = tk.Button(header, text="Logout", font=FONT_MD, bg=THEME["danger"], fg="white",
                                   command=self.app.logout, cursor="hand2")
            logout_btn.pack(side="right", padx=8, pady=8)

            # Body: Sidebar + Main
            body = tk.Frame(self, bg=THEME["bg"])
            body.pack(fill="both", expand=True)

            self.sidebar = tk.Frame(body, bg="#ECF0F1", width=200)
            self.sidebar.pack(side="left", fill="y")

            # Sidebar buttons (min height 120px for touch)
            def add_btn(text, cmd):
                b = tk.Button(self.sidebar, text=text, command=cmd, font=FONT_LG, bg=THEME["primary"], fg="white",
                              activebackground=THEME["accent"], height=3, cursor="hand2", relief="flat")
                b.pack(fill="x", padx=8, pady=6)

            add_btn("Dashboard Home", self.show_home)
            add_btn("Employees", self.show_employees)
            add_btn("Suppliers", self.show_suppliers)
            add_btn("Products", self.show_products)
            add_btn("Sales", self.show_sales)
            add_btn("Customer", self.show_customers)
            add_btn("Reports", self.show_reports)

            self.main = tk.Frame(body, bg=THEME["bg"])
            self.main.pack(side="left", fill="both", expand=True)

            self.show_home()

    def update_header_clock(self):
            self.dt_lbl.config(text=now_str())
            self.after(1000, self.update_header_clock)

    def clear_main(self):
            if self.current_section_frame:
                self.current_section_frame.destroy()
            self.current_section_frame = tk.Frame(self.main, bg=THEME["bg"])
            self.current_section_frame.pack(fill="both", expand=True)

    # --------- Sections ---------

    def show_home(self):
        self.clear_main()
        f = self.current_section_frame

        # KPI Cards
        kpi_wrap = tk.Frame(f, bg=THEME["bg"])
        kpi_wrap.pack(fill="x", padx=16, pady=16)

        def kpi(title, value, bgc):
            card = tk.Frame(kpi_wrap, bg=bgc, bd=0, relief="ridge")
            card.pack(side="left", padx=8, pady=8, fill="x", expand=True)
            tk.Label(card, text=title, font=FONT_LG, fg="white", bg=bgc).pack(anchor="w", padx=12, pady=(12, 4))
            tk.Label(card, text=value, font=("Segoe UI", 22, "bold"), fg="white", bg=bgc).pack(anchor="w", padx=12,
                                                                                               pady=(0, 12))

        con = db()
        cur = con.cursor()
        cur.execute("SELECT COUNT(*) FROM employees");
        total_emps = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM products");
        total_products = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM suppliers");
        total_suppliers = cur.fetchone()[0]
        cur.execute("SELECT IFNULL(SUM(quantity*mrp),0) FROM products");
        total_inventory_price = cur.fetchone()[0] or 0.0

        # Today's sales
        cur.execute("SELECT IFNULL(SUM(effective_total),0) FROM sales WHERE date=?", (today_str(),))
        todays_sales = cur.fetchone()[0] or 0.0

        # Low stock
        cur.execute("SELECT COUNT(*) FROM products WHERE quantity < reorder_level")
        low_stock_count = cur.fetchone()[0]
        con.close()

        kpi("Total Employees", total_emps, "#2196F3")  # blue
        kpi("Total Products", total_products, "#E53935")  # red
        kpi("Total Suppliers", total_suppliers, "#43A047")  # green
        kpi("Today’s Sales", f"₹{todays_sales:.2f}", "#FBC02D")  # yellow
        kpi("Low Stock Count", low_stock_count, THEME["accent"])

        # Admin can see online status
        if self.app.current_user[1] == "Admin":
            online_frame = tk.LabelFrame(f, text="Online Status (Admin)", bg=THEME["bg"], font=FONT_LG,
                                         fg=THEME["dark"])
            online_frame.pack(fill="x", padx=16, pady=8)
            cols = ("username", "role", "is_online", "last_login")
            tv = ttk.Treeview(online_frame, columns=cols, show="headings", height=4)
            for c in cols:
                tv.heading(c, text=c.title())
                tv.column(c, width=160)
            tv.pack(fill="x", padx=8, pady=8)
            con = db()
            cur = con.cursor()
            cur.execute(
                "SELECT username, role, is_online, IFNULL(last_login,'') last_login FROM users ORDER BY role DESC, username")
            for r in cur.fetchall():
                tv.insert("", "end",
                          values=(r["username"], r["role"], "Online" if r["is_online"] else "Offline", r["last_login"]))
            con.close()

        # Sales Graph (daily last 14 days)
        def show_graph():
            con = db()
            cur = con.cursor()
            cur.execute("""
                        SELECT date, IFNULL(SUM (effective_total), 0) AS total
                        FROM sales
                        WHERE date >= ?
                        GROUP BY date
                        ORDER BY date
                        """, ((dt.date.today() - dt.timedelta(days=13)).isoformat(),))
            data = cur.fetchall()
            con.close()
            dates = []
            totals = []
            # ensure all dates present
            for i in range(14):
                d = (dt.date.today() - dt.timedelta(days=13 - i)).isoformat()
                dates.append(d)
                total = 0.0
                for r in data:
                    if r["date"] == d:
                        total = r["total"]
                        break
                totals.append(total)
            plt.figure()
            plt.plot(dates, totals, marker="o")
            plt.title("Sales – Last 14 Days")
            plt.xlabel("Date");
            plt.ylabel("Revenue")
            plt.xticks(rotation=45, ha="right")
            plt.tight_layout()
            plt.show()

        btn = tk.Button(f, text="Show Sales Graph", font=FONT_LG, bg=THEME["primary"], fg="white",
                        cursor="hand2", command=show_graph)
        btn.pack(pady=8)

        # 🔔 NEW ALERTS BUTTON
        alerts_btn = tk.Button(f, text="Alerts (Low Stock)", font=FONT_LG,
                               bg=THEME["warning"], fg="black",
                               cursor="hand2", command=self.show_alerts)
        alerts_btn.pack(pady=8)

    def show_alerts(self):
        win = tk.Toplevel(self)
        win.title("Low Stock Alerts")
        win.geometry("500x300")
        win.configure(bg=THEME["bg"])

        cols = ("Product Name", "Quantity", "Reorder Level")
        tv = ttk.Treeview(win, columns=cols, show="headings", height=12)
        for c, w in zip(cols, [200, 100, 150]):
            tv.heading(c, text=c)
            tv.column(c, width=w, anchor="center")
        tv.pack(fill="both", expand=True, padx=10, pady=10)

        setup_treeview_striped(tv)

        # Fetch low stock products
        con = db();
        cur = con.cursor()
        cur.execute("""SELECT name, quantity, reorder_level
                       FROM products
                       WHERE quantity < reorder_level""")
        rows = cur.fetchall();
        con.close()

        for r in rows:
            tv.insert("", "end", values=(r["name"], r["quantity"], r["reorder_level"]))

    def show_employees(self):
        if self.app.current_user[1] != "Admin":
            messagebox.showwarning("Access", "Employees section is Admin only.")
            return
        self.clear_main()
        SectionEmployees(self.current_section_frame)

    def show_suppliers(self):
        if self.app.current_user[1] != "Admin":
            messagebox.showwarning("Access", "Suppliers section is Admin only.")
            return
        self.clear_main()
        SectionSuppliers(self.current_section_frame)

    def show_products(self):
        self.clear_main()
        SectionProducts(self.current_section_frame, self.app.current_user)

    def show_sales(self):
        self.clear_main()
        SectionSales(self.current_section_frame, self.app.current_user)

    def show_customers(self):
        self.clear_main()
        SectionCustomers(self.current_section_frame, self.app.current_user)

    def show_reports(self):
        username, role = self.app.current_user
        if role != "Admin":
            messagebox.showwarning("Access", "Reports section is Admin only.")
            return
        self.clear_main()
        SectionReports(self.current_section_frame, username, role)


# ---------- Section Base Utilities ----------

def setup_treeview_striped(tv: ttk.Treeview):
    style = ttk.Style()
    style.configure("Treeview", rowheight=28, font=FONT_MD)
    style.map("Treeview", background=[("selected", THEME["primary"])], foreground=[("selected", "white")])
    tv.tag_configure("odd", background="#FAFAFA")
    tv.tag_configure("even", background="#ECEFF1")


def insert_rows_striped(tv: ttk.Treeview, rows: List[Tuple[Any, ...]]):
    tv.delete(*tv.get_children())
    for i, row in enumerate(rows):
        tv.insert("", "end", values=row, tags=("even" if i % 2 == 0 else "odd",))


# ---------- Employees ----------

class SectionEmployees(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg=THEME["bg"])
        self.pack(fill="both", expand=True)

        # Search
        sframe = tk.Frame(self, bg=THEME["bg"])
        sframe.pack(fill="x", padx=12, pady=8)
        tk.Label(sframe, text="Search (Name/ID/Phone/Email):", bg=THEME["bg"], font=FONT_MD).pack(side="left")
        self.q = tk.StringVar()
        tk.Entry(sframe, textvariable=self.q, font=FONT_MD).pack(side="left", padx=8)
        tk.Button(sframe, text="Search", font=FONT_MD, bg=THEME["primary"], fg="white", command=self.refresh).pack(
            side="left", padx=4)
        tk.Button(sframe, text="Reset", font=FONT_MD, command=lambda: [self.q.set(""), self.refresh()]).pack(
            side="left", padx=4)

        # Table
        cols = ("emp_id", "name", "phone", "email", "role", "join_date")
        self.tv = ttk.Treeview(self, columns=cols, show="headings")
        for c, w in zip(cols, [80, 180, 120, 220, 120, 120]):
            self.tv.heading(c, text=c.replace("_", " ").title())
            self.tv.column(c, width=w, anchor="center")
        self.tv.pack(fill="both", expand=True, padx=12, pady=8)
        setup_treeview_striped(self.tv)

        # Form
        form = tk.LabelFrame(self, text="Add / Edit Employee", bg=THEME["bg"], font=FONT_LG, fg=THEME["dark"])
        form.pack(fill="x", padx=12, pady=8)

        self.emp_id = tk.StringVar()
        self.name = tk.StringVar()
        self.phone = tk.StringVar()
        self.email = tk.StringVar()
        self.role = tk.StringVar(value="Employee")
        self.join_date = tk.StringVar(value=today_str())

        def add_row(lbl, var, row, col, width=25):
            tk.Label(form, text=lbl, font=FONT_MD, bg=THEME["bg"]).grid(row=row, column=col * 2, padx=8, pady=6,
                                                                        sticky="e")
            tk.Entry(form, textvariable=var, font=FONT_MD, width=width).grid(row=row, column=col * 2 + 1, padx=8,
                                                                             pady=6, sticky="w")

        add_row("Emp ID", self.emp_id, 0, 0)
        add_row("Name", self.name, 0, 1)
        add_row("Phone", self.phone, 1, 0)
        add_row("Email", self.email, 1, 1)
        add_row("Role (Admin/Employee)", self.role, 2, 0)
        add_row("Join Date (YYYY-MM-DD)", self.join_date, 2, 1)

        tk.Button(form, text="Auto ID", font=FONT_MD, command=self.auto_id).grid(row=0, column=4, padx=8)
        tk.Button(form, text="Add / Save", font=FONT_MD, bg=THEME["success"], fg="white", command=self.save).grid(row=3,
                                                                                                                  column=1,
                                                                                                                  pady=8)
        tk.Button(form, text="Delete", font=FONT_MD, bg=THEME["danger"], fg="white", command=self.delete).grid(row=3,
                                                                                                               column=2,
                                                                                                               pady=8)
        tk.Button(form, text="Load Selected", font=FONT_MD, command=self.load_selected).grid(row=3, column=3, pady=8)
        tk.Button(form, text="Create/Sync User Login", font=FONT_MD, bg=THEME["accent"],
                  command=self.create_user_for_employee).grid(row=3, column=4, pady=8)

        self.refresh()

    def auto_id(self):
        self.emp_id.set(padded_id("employees", "emp_id"))

    def refresh(self):
        q = f"%{self.q.get().strip()}%"
        con = db()
        cur = con.cursor()
        cur.execute("""
                    SELECT emp_id, name, phone, email, role, join_date
                    FROM employees
                    WHERE emp_id LIKE ?
                       OR name LIKE ?
                       OR phone LIKE ?
                       OR email LIKE ?
                    ORDER BY CAST(emp_id AS INTEGER)
                    """, (q, q, q, q))
        rows = [(r["emp_id"], r["name"], r["phone"], r["email"], r["role"], r["join_date"]) for r in cur.fetchall()]
        con.close()
        insert_rows_striped(self.tv, rows)

    def save(self):
        emp_id = self.emp_id.get().strip()
        name = self.name.get().strip()
        phone = self.phone.get().strip()
        email = self.email.get().strip()
        role = self.role.get().strip()
        jdate = self.join_date.get().strip()

        if not emp_id:
            messagebox.showerror("Validation", "Emp ID required (use Auto ID).")
            return
        if not name:
            messagebox.showerror("Validation", "Name required.")
            return

        if not validate_phone(phone):
            messagebox.showerror("Validation", "Phone must be 10 digits starting 6-9 and unique.")
            return
        if not validate_email(email):
            messagebox.showerror("Validation", "Email must be @gmail.com or @yahoo.com and unique.")
            return
        try:
            d = dt.date.fromisoformat(jdate)
            if d > dt.date.today():
                messagebox.showerror("Validation", "Join date cannot exceed system date.")
                return
        except:
            messagebox.showerror("Validation", "Join date must be YYYY-MM-DD.")
            return
        if role not in ("Admin", "Employee"):
            messagebox.showerror("Validation", "Role must be Admin or Employee.")
            return

        con = db()
        cur = con.cursor()
        # upsert
        try:
            cur.execute("INSERT INTO employees(emp_id,name,phone,email,role,join_date) VALUES(?,?,?,?,?,?)",
                        (emp_id, name, phone, email, role, jdate))
            con.commit()
        except sqlite3.IntegrityError:
            # update existing
            cur.execute("""UPDATE employees
                           SET name=?,
                               phone=?,
                               email=?,
                               role=?,
                               join_date=?
                           WHERE emp_id = ?""",
                        (name, phone, email, role, jdate, emp_id))
            con.commit()
        con.close()
        messagebox.showinfo("Saved", "Employee saved.")
        self.refresh()

    def delete(self):
        sel = self.tv.selection()
        if not sel:
            messagebox.showwarning("Delete", "Select a row to delete.")
            return
        emp_id = self.tv.item(sel[0], "values")[0]
        if not messagebox.askyesno("Confirm", f"Delete employee {emp_id}?"):
            return
        con = db()
        cur = con.cursor()
        cur.execute("DELETE FROM employees WHERE emp_id=?", (emp_id,))
        con.commit()
        con.close()
        self.refresh()

    def load_selected(self):
        sel = self.tv.selection()
        if not sel:
            return
        vals = self.tv.item(sel[0], "values")
        self.emp_id.set(vals[0]);
        self.name.set(vals[1]);
        self.phone.set(vals[2])
        self.email.set(vals[3]);
        self.role.set(vals[4]);
        self.join_date.set(vals[5])

    def create_user_for_employee(self):
        # Create/Update a user with username=emp_id or employee name? We'll use employee name (no spaces, lower).
        name = self.name.get().strip()
        role = self.role.get().strip()
        if not name:
            messagebox.showerror("User", "Load or enter an employee first.")
            return
        username = re.sub(r"\s+", "", name).lower()
        password = employee_default_password(name)
        con = db()
        cur = con.cursor()
        try:
            cur.execute("INSERT INTO users(username,password,role,is_online,last_login) VALUES(?,?,?,?,?)",
                        (username, password, role, 0, None))
        except sqlite3.IntegrityError:
            cur.execute("UPDATE users SET password=?, role=? WHERE username=?", (password, role, username))
        con.commit()
        con.close()
        messagebox.showinfo("User", f"User created/updated.\nUsername: {username}\nPassword: {password}")


# ---------- Suppliers ----------

class SectionSuppliers(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg=THEME["bg"])
        self.pack(fill="both", expand=True)

        sframe = tk.Frame(self, bg=THEME["bg"])
        sframe.pack(fill="x", padx=12, pady=8)
        tk.Label(sframe, text="Search (Name/Contact/Company):", bg=THEME["bg"], font=FONT_MD).pack(side="left")
        self.q = tk.StringVar()
        tk.Entry(sframe, textvariable=self.q, font=FONT_MD).pack(side="left", padx=8)
        tk.Button(sframe, text="Search", font=FONT_MD, bg=THEME["primary"], fg="white", command=self.refresh).pack(
            side="left", padx=4)
        tk.Button(sframe, text="Reset", font=FONT_MD, command=lambda: [self.q.set(""), self.refresh()]).pack(
            side="left", padx=4)

        cols = ("supplier_id", "name", "company", "phone", "email", "address")
        self.tv = ttk.Treeview(self, columns=cols, show="headings")
        for c, w in zip(cols, [80, 150, 150, 120, 220, 220]):
            self.tv.heading(c, text=c.replace("_", " ").title())
            self.tv.column(c, width=w, anchor="center")
        self.tv.pack(fill="both", expand=True, padx=12, pady=8)
        setup_treeview_striped(self.tv)

        form = tk.LabelFrame(self, text="Add / Edit Supplier", bg=THEME["bg"], font=FONT_LG, fg=THEME["dark"])
        form.pack(fill="x", padx=12, pady=8)

        self.supplier_id = tk.StringVar()
        self.name = tk.StringVar()
        self.company = tk.StringVar()
        self.phone = tk.StringVar()
        self.email = tk.StringVar()
        self.address = tk.StringVar()

        def add(lbl, var, r, c, w=25):
            tk.Label(form, text=lbl, font=FONT_MD, bg=THEME["bg"]).grid(row=r, column=c * 2, padx=8, pady=6, sticky="e")
            tk.Entry(form, textvariable=var, font=FONT_MD, width=w).grid(row=r, column=c * 2 + 1, padx=8, pady=6,
                                                                         sticky="w")

        add("Supplier ID", self.supplier_id, 0, 0)
        add("Name", self.name, 0, 1)
        add("Company", self.company, 1, 0)
        add("Phone", self.phone, 1, 1)
        add("Email", self.email, 2, 0)
        add("Address", self.address, 2, 1, 40)

        tk.Button(form, text="Auto ID", font=FONT_MD, command=self.auto_id).grid(row=0, column=4, padx=8)
        tk.Button(form, text="Add / Save", font=FONT_MD, bg=THEME["success"], fg="white", command=self.save).grid(row=3,
                                                                                                                  column=1,
                                                                                                                  pady=8)
        tk.Button(form, text="Delete", font=FONT_MD, bg=THEME["danger"], fg="white", command=self.delete).grid(row=3,
                                                                                                               column=2,
                                                                                                               pady=8)
        tk.Button(form, text="Load Selected", font=FONT_MD, command=self.load_selected).grid(row=3, column=3, pady=8)

        self.refresh()

    def auto_id(self):
        self.supplier_id.set(padded_id("suppliers", "supplier_id"))

    def refresh(self):
        q = f"%{self.q.get().strip()}%"
        con = db()
        cur = con.cursor()
        cur.execute("""
                    SELECT supplier_id, name, company, phone, email, address
                    FROM suppliers
                    WHERE name LIKE ?
                       OR phone LIKE ?
                       OR company LIKE ?
                    ORDER BY CAST(supplier_id AS INTEGER)
                    """, (q, q, q))
        rows = [(r["supplier_id"], r["name"], r["company"], r["phone"], r["email"], r["address"]) for r in
                cur.fetchall()]
        con.close()
        insert_rows_striped(self.tv, rows)

    def save(self):
        sid = self.supplier_id.get().strip()
        name = self.name.get().strip()
        company = self.company.get().strip()
        phone = self.phone.get().strip()
        email = self.email.get().strip()
        address = self.address.get().strip()
        if not sid: messagebox.showerror("Validation", "Supplier ID required (Auto ID)."); return
        if not name or not company: messagebox.showerror("Validation", "Name and Company are required."); return
        if not validate_phone(phone): messagebox.showerror("Validation", "Phone invalid/duplicate."); return
        if not validate_email(email): messagebox.showerror("Validation",
                                                           "Email must be @gmail.com or @yahoo.com."); return
        con = db()
        cur = con.cursor()
        try:
            cur.execute("INSERT INTO suppliers(supplier_id,name,company,phone,email,address) VALUES(?,?,?,?,?,?)",
                        (sid, name, company, phone, email, address))
            con.commit()
        except sqlite3.IntegrityError:
            cur.execute("""UPDATE suppliers
                           SET name=?,
                               company=?,
                               phone=?,
                               email=?,
                               address=?
                           WHERE supplier_id = ?""",
                        (name, company, phone, email, address, sid))
            con.commit()
        con.close()
        messagebox.showinfo("Saved", "Supplier saved.")
        self.refresh()

    def delete(self):
        sel = self.tv.selection()
        if not sel:
            messagebox.showwarning("Delete", "Select a row.")
            return
        sid = self.tv.item(sel[0], "values")[0]
        if not messagebox.askyesno("Confirm", f"Delete supplier {sid}?"):
            return
        con = db()
        cur = con.cursor()
        cur.execute("DELETE FROM suppliers WHERE supplier_id=?", (sid,))
        con.commit()
        con.close()
        self.refresh()

    def load_selected(self):
        sel = self.tv.selection()
        if not sel: return
        v = self.tv.item(sel[0], "values")
        self.supplier_id.set(v[0]);
        self.name.set(v[1]);
        self.company.set(v[2])
        self.phone.set(v[3]);
        self.email.set(v[4]);
        self.address.set(v[5])


# ---------- Products ----------

# ---------- Products ----------

class SectionProducts(tk.Frame):
    def __init__(self, parent, user):
        super().__init__(parent, bg=THEME["bg"])
        self.pack(fill="both", expand=True)

        # Ensure GST column exists in DB
        self.ensure_gst_column()

        # ================= TOP BAR =================
        top = tk.Frame(self, bg=THEME["bg"])
        top.pack(fill="x", padx=12, pady=8)
        tk.Label(top, text="Search (Name/Category/Company):", bg=THEME["bg"], font=FONT_MD).pack(side="left")
        self.q = tk.StringVar()
        tk.Entry(top, textvariable=self.q, font=FONT_MD).pack(side="left", padx=8)
        tk.Button(top, text="Search", font=FONT_MD, bg=THEME["primary"], fg="white",
                  command=self.refresh).pack(side="left", padx=4)
        tk.Button(top, text="Reset", font=FONT_MD,
                  command=lambda: [self.q.set(""), self.refresh()]).pack(side="left", padx=4)
        tk.Button(top, text="Export Excel", font=FONT_MD,
                  command=lambda: self.export_excel()).pack(side="right", padx=4)
        tk.Button(top, text="Export PDF", font=FONT_MD,
                  command=lambda: self.export_pdf()).pack(side="right", padx=4)

        self.total_lbl = tk.Label(top, text="Total Inventory Price: ₹0.00",
                                  bg=THEME["bg"], font=FONT_LG, fg=THEME["dark"])
        self.total_lbl.pack(side="right", padx=12)

        # ================= TABLE =================
        cols = ("product_id", "name", "category", "supplier_id", "supplier_company",
                "quantity", "unit_price", "gst", "mrp", "reorder_level", "Low_Stock")
        self.tv = ttk.Treeview(self, columns=cols, show="headings")
        widths = [80, 200, 150, 100, 180, 80, 100, 60, 100, 120, 100]
        for c, w in zip(cols, widths):
            self.tv.heading(c, text=c.replace("_", " "))
            self.tv.column(c, width=w, anchor="center")
        self.tv.pack(fill="both", expand=True, padx=12, pady=8)
        setup_treeview_striped(self.tv)

        # ================= FORM =================
        form = tk.LabelFrame(self, text="Add / Edit Product",
                             bg=THEME["bg"], font=FONT_LG, fg=THEME["dark"])
        form.pack(fill="x", padx=12, pady=8)

        self.product_id = tk.StringVar()
        self.name = tk.StringVar()
        self.category = tk.StringVar()
        self.supplier_id = tk.StringVar()
        self.quantity = tk.StringVar(value="0")
        self.unit_price = tk.StringVar(value="0.00")
        self.gst = tk.StringVar(value="18")  # default GST
        self.mrp = tk.StringVar(value="0.00")
        self.reorder_level = tk.StringVar(value="0")

        # Auto-calc MRP from Unit Price + GST
        def update_mrp(*args):
            try:
                u = float(self.unit_price.get())
                g = float(self.gst.get())
                if u < 0:
                    self.unit_price.set("0.00")
                    u = 0.0
                g = min(max(g, 0), 40)
                self.gst.set(str(int(g)))
                self.mrp.set(f"{u * (1 + g / 100):.2f}")
            except:
                self.mrp.set("0.00")

        self.unit_price.trace("w", update_mrp)
        self.gst.trace("w", update_mrp)

        def add(lbl, var, r, c, w=25, ro=False):
            tk.Label(form, text=lbl, font=FONT_MD, bg=THEME["bg"]).grid(
                row=r, column=c * 2, padx=8, pady=6, sticky="e")
            state = "readonly" if ro else "normal"
            tk.Entry(form, textvariable=var, font=FONT_MD, width=w, state=state).grid(
                row=r, column=c * 2 + 1, padx=8, pady=6, sticky="w")

        # --- Row 0 ---
        add("SKU", self.product_id, 0, 0)
        tk.Button(form, text="Auto SKU", font=FONT_MD, command=self.auto_id).grid(
            row=0, column=2, padx=8, pady=6, sticky="w")
        add("Name", self.name, 0, 3)

        # --- Row 1 ---
        add("Category", self.category, 1, 0)
        tk.Label(form, text="Supplier", font=FONT_MD, bg=THEME["bg"]).grid(
            row=1, column=2, padx=8, pady=6, sticky="e")
        self.supplier_cmb = ttk.Combobox(form, textvariable=self.supplier_id, width=27, state="readonly")
        self.supplier_cmb.grid(row=1, column=3, padx=8, pady=6, sticky="w")
        tk.Label(form, text="GST (%)", font=FONT_MD, bg=THEME["bg"]).grid(
            row=1, column=4, padx=8, pady=6, sticky="e")
        self.gst_spin = tk.Spinbox(form, from_=0, to=40, increment=1, textvariable=self.gst,
                                   font=FONT_MD, width=5, state="readonly")
        self.gst_spin.grid(row=1, column=5, padx=8, pady=6, sticky="w")

        # --- Row 2 ---
        add("Quantity", self.quantity, 2, 0)
        add("Unit Price", self.unit_price, 2, 1)
        add("MRP (Auto)", self.mrp, 2, 2, ro=True)
        add("Reorder Level", self.reorder_level, 2, 3)

        # --- Action Buttons (Row 3, spanning across) ---
        btn_frame = tk.Frame(form, bg=THEME["bg"])
        btn_frame.grid(row=3, column=0, columnspan=6, pady=10)

        tk.Button(btn_frame, text="Add / Save", font=FONT_MD, bg=THEME["success"], fg="white",
                  command=self.save, width=12).pack(side="left", padx=10)
        tk.Button(btn_frame, text="Delete", font=FONT_MD, bg=THEME["danger"], fg="white",
                  command=self.delete, width=12).pack(side="left", padx=10)
        tk.Button(btn_frame, text="Load Selected", font=FONT_MD,
                  command=self.load_selected, width=15).pack(side="left", padx=10)

        # --- Load initial data ---
        self.load_suppliers()
        self.refresh()


    # ================= DB MIGRATION =================
    def ensure_gst_column(self):
        con = db()
        cur = con.cursor()
        cur.execute("PRAGMA table_info(products)")
        cols = [row[1] for row in cur.fetchall()]
        if "gst" not in cols:
            cur.execute("ALTER TABLE products ADD COLUMN gst REAL DEFAULT 18")
            con.commit()
        con.close()

    # ================= SUPPLIERS =================
    def load_suppliers(self):
        con = db()
        cur = con.cursor()
        cur.execute("SELECT supplier_id, company FROM suppliers ORDER BY company")
        self.suppliers = cur.fetchall()
        con.close()
        self.supplier_cmb["values"] = [f"{r['supplier_id']} - {r['company']}" for r in self.suppliers]

    def auto_id(self):
        self.product_id.set(padded_id("products", "product_id"))

    # ================= REFRESH =================
    def refresh(self):
        q = f"%{self.q.get().strip()}%"
        con = db()
        cur = con.cursor()
        cur.execute("""
                    SELECT p.product_id,
                           p.name,
                           p.category,
                           p.supplier_id,
                           s.company,
                           p.quantity,
                           p.unit_price,
                           p.gst,
                           p.mrp,
                           p.reorder_level,
                           CASE WHEN p.quantity < p.reorder_level THEN 'YES' ELSE 'NO' END AS low_stock
                    FROM products p
                             JOIN suppliers s ON s.supplier_id = p.supplier_id
                    WHERE p.name LIKE ?
                       OR p.category LIKE ?
                       OR s.company LIKE ?
                    ORDER BY CAST(p.product_id AS INTEGER)
                    """, (q, q, q))
        rows = [(r["product_id"], r["name"], r["category"], r["supplier_id"], r["company"],
                 r["quantity"], f"{r['unit_price']:.2f}", f"{r['gst']:.0f}%", f"{r['mrp']:.2f}",
                 r["reorder_level"], r["low_stock"]) for r in cur.fetchall()]
        cur.execute("SELECT IFNULL(SUM(quantity*unit_price),0) FROM products")
        total_val = cur.fetchone()[0] or 0.0
        con.close()
        insert_rows_striped(self.tv, rows)
        self.total_lbl.config(text=f"Total Inventory Price: ₹{total_val:.2f}")

    # ================= SAVE =================
    def save(self):
        pid = self.product_id.get().strip()
        name = self.name.get().strip()
        cat = self.category.get().strip()
        supplier_text = self.supplier_id.get().strip()
        qty = self.quantity.get().strip()
        unit_price = self.unit_price.get().strip()
        gst = self.gst.get().strip()
        rl = self.reorder_level.get().strip()

        if not pid:
            messagebox.showerror("Validation", "SKU required (Auto SKU).")
            return
        if not name or not cat:
            messagebox.showerror("Validation", "Name and Category required.")
            return
        if "-" not in supplier_text:
            messagebox.showerror("Validation", "Select supplier from dropdown.")
            return
        supplier_id = supplier_text.split(" - ")[0].strip()

        try:
            qty = int(qty)
            rl = int(rl)
            unit_price = float(unit_price)
            gst = float(gst)
            gst = min(max(gst, 0), 40)
            mrp = round(unit_price * (1 + gst / 100), 2)
        except:
            messagebox.showerror("Validation", "Invalid numbers in Quantity/Price/GST/Reorder.")
            return

        if qty < 0 or rl < 0 or unit_price < 0:
            messagebox.showerror("Validation", "Negative values not allowed.")
            return

        con = db()
        cur = con.cursor()
        try:
            cur.execute("""INSERT INTO products(product_id, name, category, supplier_id,
                                                quantity, unit_price, gst, mrp, reorder_level)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        (pid, name, cat, supplier_id, qty, unit_price, gst, mrp, rl))
            con.commit()
        except sqlite3.IntegrityError:
            cur.execute("""UPDATE products
                           SET name=?,
                               category=?,
                               supplier_id=?,
                               quantity=?,
                               unit_price=?,
                               gst=?,
                               mrp=?,
                               reorder_level=?
                           WHERE product_id = ?""",
                        (name, cat, supplier_id, qty, unit_price, gst, mrp, rl, pid))
            con.commit()
        con.close()
        messagebox.showinfo("Saved", "Product saved.")
        self.refresh()

    # ================= DELETE =================
    def delete(self):
        sel = self.tv.selection()
        if not sel:
            messagebox.showwarning("Delete", "Select a row.")
            return
        pid = self.tv.item(sel[0], "values")[0]
        if not messagebox.askyesno("Confirm", f"Delete product {pid}?"):
            return
        con = db()
        cur = con.cursor()
        cur.execute("DELETE FROM products WHERE product_id=?", (pid,))
        con.commit()
        con.close()
        self.refresh()

    # ================= LOAD SELECTED =================
    def load_selected(self):
        sel = self.tv.selection()
        if not sel:
            return
        v = self.tv.item(sel[0], "values")
        self.product_id.set(v[0])
        self.name.set(v[1])
        self.category.set(v[2])
        sid = v[3]
        for i, r in enumerate(self.suppliers):
            if r["supplier_id"] == sid:
                self.supplier_cmb.current(i)
                break
        self.quantity.set(v[5])
        self.unit_price.set(v[6])
        self.gst.set(v[7].replace("%", ""))
        self.mrp.set(v[8])
        self.reorder_level.set(v[9])

    # ================= EXPORT =================
    def export_excel(self):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([self.tv.heading(c)["text"] for c in self.tv["columns"]])
        for row_id in self.tv.get_children():
            row = list(self.tv.item(row_id)["values"])
            gst_index = list(self.tv["columns"]).index("gst")
            row[gst_index] = row[gst_index].replace("%", "")
            ws.append(row)
        wb.save("products.xlsx")
        messagebox.showinfo("Export", "Products exported to products.xlsx")

    def export_pdf(self):
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet

        doc = SimpleDocTemplate("products.pdf", pagesize=A4)
        elements = []
        style = getSampleStyleSheet()

        data = [[self.tv.heading(c)["text"] for c in self.tv["columns"]]]
        for row_id in self.tv.get_children():
            row = list(self.tv.item(row_id)["values"])
            gst_index = list(self.tv["columns"]).index("gst")
            row[gst_index] = row[gst_index].replace("%", "")
            data.append(row)

        table = Table(data)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(Paragraph("Products", style["Title"]))
        elements.append(table)
        doc.build(elements)
        messagebox.showinfo("Export", "Products exported to products.pdf")


# ---------- Customers ----------

class SectionCustomers(tk.Frame):
    def __init__(self, parent, user):
        super().__init__(parent, bg=THEME["bg"])
        self.pack(fill="both", expand=True)

        # ---------------- Search ----------------
        sframe = tk.Frame(self, bg=THEME["bg"])
        sframe.pack(fill="x", padx=12, pady=8)
        tk.Label(sframe, text="Search (Name/Contact):", bg=THEME["bg"], font=FONT_MD).pack(side="left")
        self.q = tk.StringVar()
        tk.Entry(sframe, textvariable=self.q, font=FONT_MD).pack(side="left", padx=8)
        tk.Button(sframe, text="Search", font=FONT_MD, bg=THEME["primary"], fg="white",
                  command=self.refresh).pack(side="left", padx=4)
        tk.Button(sframe, text="Reset", font=FONT_MD,
                  command=lambda: [self.q.set(""), self.refresh()]).pack(side="left", padx=4)

        # --- Bulk Communication Button ---
        tk.Button(sframe, text="Bulk Mail / SMS", font=FONT_MD, bg=THEME["accent"], fg="white",
                  command=self.bulk_comm_window).pack(side="right", padx=6)

        # ---------------- Table ----------------
        cols = ("customer_id", "name", "phone", "email")
        self.tv = ttk.Treeview(self, columns=cols, show="headings")
        for c, w in zip(cols, [80, 180, 120, 220]):
            self.tv.heading(c, text=c.replace("_", " ").title())
            self.tv.column(c, width=w, anchor="center")
        self.tv.pack(fill="both", expand=True, padx=12, pady=8)
        setup_treeview_striped(self.tv)

        # ---------------- Form ----------------
        form = tk.LabelFrame(self, text="Add / Edit Customer", bg=THEME["bg"],
                             font=FONT_LG, fg=THEME["dark"])
        form.pack(fill="x", padx=12, pady=8)

        self.customer_id = tk.StringVar()
        self.name = tk.StringVar()
        self.phone = tk.StringVar()
        self.email = tk.StringVar()

        def add(lbl, var, r, c, w=25):
            tk.Label(form, text=lbl, font=FONT_MD, bg=THEME["bg"]).grid(
                row=r, column=c * 2, padx=8, pady=6, sticky="e")
            tk.Entry(form, textvariable=var, font=FONT_MD, width=w).grid(
                row=r, column=c * 2 + 1, padx=8, pady=6, sticky="w")

        add("Customer ID", self.customer_id, 0, 0)
        add("Name", self.name, 0, 1)
        add("Phone", self.phone, 1, 0)
        add("Email", self.email, 1, 1)

        tk.Button(form, text="Auto ID", font=FONT_MD,
                  command=self.auto_id).grid(row=0, column=4, padx=8)
        tk.Button(form, text="Add / Save", font=FONT_MD, bg=THEME["success"], fg="white",
                  command=self.save).grid(row=3, column=1, pady=8)
        tk.Button(form, text="Delete", font=FONT_MD, bg=THEME["danger"], fg="white",
                  command=self.delete).grid(row=3, column=2, pady=8)
        tk.Button(form, text="Load Selected", font=FONT_MD,
                  command=self.load_selected).grid(row=3, column=3, pady=8)

        self.refresh()

    # ---------------- BASIC FUNCTIONS ----------------
    def auto_id(self):
        self.customer_id.set(padded_id("customers", "customer_id"))

    def refresh(self):
        q = f"%{self.q.get().strip()}%"
        con = db()
        cur = con.cursor()
        cur.execute("""
                    SELECT customer_id, name, phone, email
                    FROM customers
                    WHERE name LIKE ?
                       OR phone LIKE ?
                    ORDER BY CAST(customer_id AS INTEGER)
                    """, (q, q))
        rows = [(r["customer_id"], r["name"], r["phone"], r["email"]) for r in cur.fetchall()]
        con.close()
        insert_rows_striped(self.tv, rows)

    def save(self):
        cid = self.customer_id.get().strip()
        name = self.name.get().strip()
        phone = self.phone.get().strip()
        email = self.email.get().strip()

        # ---------------- Validation ----------------
        if not cid:
            messagebox.showerror("Validation", "Customer ID required (Auto ID).")
            return
        if not name:
            messagebox.showerror("Validation", "Name required.")
            return
        if not re.match(r'^[A-Za-z ]+$', name):  # alphabets + spaces only
            messagebox.showerror("Validation", "Name must contain only alphabets and spaces.")
            return
        if not validate_phone(phone):
            messagebox.showerror("Validation", "Phone invalid/duplicate.")
            return
        if not validate_email(email):
            messagebox.showerror("Validation", "Email must be @gmail.com or @yahoo.com.")
            return

        # ---------------- Insert / Update ----------------
        con = db()
        cur = con.cursor()
        try:
            cur.execute("INSERT INTO customers(customer_id,name,phone,email) VALUES(?,?,?,?)",
                        (cid, name, phone, email))
            con.commit()
        except sqlite3.IntegrityError:
            cur.execute("""UPDATE customers
                           SET name=?, phone=?, email=?
                           WHERE customer_id = ?""",
                        (name, phone, email, cid))
            con.commit()
        con.close()
        messagebox.showinfo("Saved", "Customer saved.")
        self.refresh()

    def delete(self):
        sel = self.tv.selection()
        if not sel:
            messagebox.showwarning("Delete", "Select a row.")
            return
        cid = self.tv.item(sel[0], "values")[0]
        if not messagebox.askyesno("Confirm", f"Delete customer {cid}?"):
            return
        con = db()
        cur = con.cursor()
        cur.execute("DELETE FROM customers WHERE customer_id=?", (cid,))
        con.commit()
        con.close()
        self.refresh()

    def load_selected(self):
        sel = self.tv.selection()
        if not sel:
            return
        v = self.tv.item(sel[0], "values")
        self.customer_id.set(v[0])
        self.name.set(v[1])
        self.phone.set(v[2])
        self.email.set(v[3])

    # ---------------- BULK COMMUNICATION ----------------
    def bulk_comm_window(self):
        win = tk.Toplevel(self)
        win.title("Bulk Mail / SMS")
        win.geometry("450x400")

        # Subject (email only)
        tk.Label(win, text="Subject (Email Only):", font=FONT_MD).pack(pady=5)
        subject_var = tk.StringVar()
        tk.Entry(win, textvariable=subject_var, width=50).pack()

        # Message (both mail + sms)
        tk.Label(win, text="Message:", font=FONT_MD).pack(pady=5)
        body_text = tk.Text(win, height=10, width=50)
        body_text.pack()

        # Mode selection
        mode_var = tk.StringVar(value="both")
        tk.Label(win, text="Send via:", font=FONT_MD).pack(pady=5)
        tk.Radiobutton(win, text="Email Only", variable=mode_var, value="email").pack()
        tk.Radiobutton(win, text="SMS Only", variable=mode_var, value="sms").pack()
        tk.Radiobutton(win, text="Both", variable=mode_var, value="both").pack()

        def send_action():
            subject = subject_var.get().strip()
            message = body_text.get("1.0", "end-1c").strip()
            mode = mode_var.get()

            if not message:
                messagebox.showwarning("Bulk Mail/SMS", "Message cannot be empty.")
                return

            con = db()
            cur = con.cursor()
            cur.execute("SELECT email, phone FROM customers")
            customers = cur.fetchall()
            con.close()

            recipients_mail = [c["email"] for c in customers if c["email"]]
            recipients_sms = [c["phone"] for c in customers if c["phone"]]

            if mode in ("email", "both") and recipients_mail:
                self.send_bulk_mail(subject or "Notification", message,
                                    "lalbaghenterprises", "lojn yuaa tcfn rqxa", recipients_mail)

            if mode in ("sms", "both") and recipients_sms:
                self.send_bulk_sms(message, recipients_sms)

        tk.Button(win, text="Send", bg="green", fg="white",
                  font=FONT_MD, command=send_action).pack(pady=15)

    def send_bulk_mail(self, subject, body, sender_email, sender_password, recipients):
        try:
            server = smtplib.SMTP("smtp.gmail.com", 587)
            server.starttls()
            server.login(sender_email, sender_password)

            for email in recipients:
                msg = MIMEMultipart()
                msg["From"] = sender_email
                msg["To"] = email
                msg["Subject"] = subject
                msg.attach(MIMEText(body, "plain"))
                server.sendmail(sender_email, email, msg.as_string())

            server.quit()
            messagebox.showinfo("Bulk Mail", "✅ Bulk mail sent successfully!")

        except Exception as e:
            messagebox.showerror("Error", f"❌ Failed to send mail:\n{e}")


# ---------- Sales ----------
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.barcode import qr

from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.barcode import qr


class SectionSales(tk.Frame):
    def __init__(self, parent, user):
        super().__init__(parent, bg=THEME["bg"])
        self.pack(fill="both", expand=True)
        self.username, self.role = user

        # ================= NEW SALE FORM =================
        form = tk.LabelFrame(self, text="New Sale / Cart", bg=THEME["bg"], font=FONT_LG, fg=THEME["dark"])
        form.pack(fill="x", padx=12, pady=8)

        # --- Product Fields ---
        self.product_pid = tk.StringVar()
        self.product_name = tk.StringVar()
        self.product_cat = tk.StringVar()
        self.product_mrp = tk.StringVar()
        self.qty = tk.StringVar(value="1")

        tk.Label(form, text="Product (ID - Name)", font=FONT_MD, bg=THEME["bg"]).grid(row=0, column=0, padx=8, pady=6, sticky="e")
        self.product_cmb = ttk.Combobox(form, textvariable=self.product_pid, width=40, state="readonly")
        self.product_cmb.grid(row=0, column=1, padx=8, pady=6, sticky="w")
        self.product_cmb.bind("<<ComboboxSelected>>", self.on_product_selected)

        def add_ro(label, var, r, c):
            tk.Label(form, text=label, font=FONT_MD, bg=THEME["bg"]).grid(row=r, column=c * 2, padx=8, pady=6, sticky="e")
            tk.Entry(form, textvariable=var, font=FONT_MD, state="readonly", width=20).grid(row=r, column=c * 2 + 1, padx=8, pady=6, sticky="w")

        add_ro("Name", self.product_name, 1, 0)
        add_ro("Category", self.product_cat, 1, 1)
        add_ro("MRP", self.product_mrp, 1, 2)

        # --- Quantity ---
        tk.Label(form, text="Quantity", font=FONT_MD, bg=THEME["bg"]).grid(row=2, column=0, padx=8, pady=6, sticky="e")
        tk.Entry(form, textvariable=self.qty, font=FONT_MD, width=8).grid(row=2, column=1, padx=8, pady=6, sticky="w")

        # --- Per Product Discount ---
        self.prod_discount_type = tk.StringVar(value="Flat")
        self.prod_discount_value = tk.StringVar(value="0")
        tk.Label(form, text="Discount", font=FONT_MD, bg=THEME["bg"]).grid(row=2, column=2, padx=8, pady=6, sticky="e")
        ttk.Combobox(form, textvariable=self.prod_discount_type, values=["Flat", "Percent"], state="readonly", width=10).grid(row=2, column=3, padx=8, pady=6, sticky="w")
        tk.Entry(form, textvariable=self.prod_discount_value, font=FONT_MD, width=10).grid(row=2, column=4, padx=8, pady=6, sticky="w")

        # --- Customer ---
        self.customer_sel = tk.StringVar()
        self.new_customer_name = tk.StringVar()
        self.new_customer_phone = tk.StringVar()
        self.new_customer_email = tk.StringVar()
        self.new_customer_address = tk.StringVar()

        tk.Label(form, text="Customer (ID - Name)", font=FONT_MD, bg=THEME["bg"]).grid(row=3, column=0, padx=8, pady=6,
                                                                                       sticky="e")
        self.customer_cmb = ttk.Combobox(form, textvariable=self.customer_sel, width=40, state="readonly")
        self.customer_cmb.grid(row=3, column=1, padx=8, pady=6, sticky="w")

        tk.Label(form, text="Or Add New", font=FONT_MD, bg=THEME["bg"]).grid(row=4, column=0, padx=8, pady=6,
                                                                             sticky="e")
        tk.Entry(form, textvariable=self.new_customer_name, font=FONT_MD, width=18).grid(row=4, column=1, padx=8,
                                                                                         pady=6, sticky="w")
        tk.Entry(form, textvariable=self.new_customer_phone, font=FONT_MD, width=14).grid(row=4, column=2, padx=8,
                                                                                          pady=6, sticky="w")
        tk.Entry(form, textvariable=self.new_customer_email, font=FONT_MD, width=20).grid(row=4, column=3, padx=8,
                                                                                          pady=6, sticky="w")


        # --- Cart Buttons ---
        tk.Button(form, text="Add to Cart", bg=THEME["accent"], fg="white", font=FONT_MD, command=self.add_to_cart).grid(row=5, column=0, padx=8, pady=8, sticky="ew")
        tk.Button(form, text="Remove Selected", bg="#D84315", fg="white", font=FONT_MD, command=self.remove_selected_from_cart).grid(row=5, column=1, padx=8, pady=8, sticky="ew")
        tk.Button(form, text="Clear Cart", bg="#9E9E9E", fg="white", font=FONT_MD, command=self.clear_cart).grid(row=5, column=2, padx=8, pady=8, sticky="ew")
        tk.Button(form, text="Checkout + Invoice", bg=THEME["success"], fg="white", font=FONT_LG, command=self.checkout).grid(row=6, column=0, columnspan=6, sticky="ew", padx=8, pady=10)
        tk.Button(form, text="Return / Refund", bg=THEME["warning"], fg="white", font=FONT_LG,
                  command=self.show_returns).grid(row=7, column=0, columnspan=6, sticky="ew", padx=8, pady=10)

        # ================= CART TABLE =================
        cart_frame = tk.LabelFrame(self, text="Cart", bg=THEME["bg"], font=FONT_LG, fg=THEME["dark"])
        cart_frame.pack(fill="both", padx=12, pady=8)

        cols = ("product_id", "product_name", "category", "qty", "mrp", "discount_type", "discount_value", "final_total")
        self.cart_tv = ttk.Treeview(cart_frame, columns=cols, show="headings", height=7)
        for c, w in zip(cols, [80, 200, 120, 60, 80, 100, 100, 100]):
            self.cart_tv.heading(c, text=c.replace("_", " ").title())
            self.cart_tv.column(c, width=w, anchor="center")
        self.cart_tv.pack(fill="both", expand=True, padx=8, pady=8)
        setup_treeview_striped(self.cart_tv)

        totals_frame = tk.Frame(cart_frame, bg=THEME["bg"])
        totals_frame.pack(fill="x", padx=8, pady=6)
        self.grand_total_var = tk.StringVar(value="₹0.00")
        tk.Label(totals_frame, text="Grand Total:", bg=THEME["bg"], font=FONT_LG).pack(side="left", padx=6)
        tk.Label(totals_frame, textvariable=self.grand_total_var, bg=THEME["bg"], font=FONT_LG).pack(side="left", padx=6)

        self.cart = []

        # ================= SALES HISTORY =================
        bot = tk.Frame(self, bg=THEME["bg"])
        bot.pack(fill="both", expand=True, padx=12, pady=8)

        tk.Label(bot, text="Sales History", bg=THEME["bg"], font=FONT_LG, fg=THEME["dark"]).pack(anchor="w", padx=8, pady=4)

        cols = ("sale_id", "date", "product_name", "category", "quantity", "mrp", "discount_type", "discount_value", "effective_total", "sold_by", "customer_name")
        self.tv = ttk.Treeview(bot, columns=cols, show="headings")
        for c, w in zip(cols, [60, 90, 200, 120, 70, 70, 80, 80, 100, 100, 160]):
            self.tv.heading(c, text=c.replace("_", " ").title())
            self.tv.column(c, width=w, anchor="center")
        self.tv.pack(fill="both", expand=True, pady=8)
        setup_treeview_striped(self.tv)

        # Load initial data
        self.load_products()
        self.load_customers()
        self.refresh()

    # ---------------- Helpers ----------------
    def has_new_customer_input(self):
        return any(v.get().strip() for v in (
            self.new_customer_name, self.new_customer_phone, self.new_customer_email, self.new_customer_address
        ))

    def _on_new_customer_typing(self, e=None):
        if self.has_new_customer_input():
            self.customer_cmb.set("")
            self.customer_sel.set("")

    # ---------------- Loaders ----------------
    def load_products(self):
        con = db(); cur = con.cursor()
        cur.execute("SELECT product_id,name,category,mrp,quantity FROM products ORDER BY name")
        self.products = cur.fetchall(); con.close()
        self.product_cmb["values"] = [f"{r['product_id']} - {r['name']}" for r in self.products]

    def on_product_selected(self, e=None):
        pid = self.product_pid.get().split(" - ")[0]
        for r in self.products:
            if r["product_id"] == pid:
                self.product_name.set(r["name"])
                self.product_cat.set(r["category"])
                self.product_mrp.set(f"{r['mrp']:.2f}")

    def load_customers(self):
        con = db(); cur = con.cursor()
        cur.execute("SELECT customer_id,name FROM customers ORDER BY name")
        self.customers = cur.fetchall(); con.close()
        self.customer_cmb["values"] = [f"{r['customer_id']} - {r['name']}" for r in self.customers]

    # ---------------- Cart Operations ----------------
    def add_to_cart(self):
        if not self.product_pid.get():
            return
        pid = self.product_pid.get().split(" - ")[0]
        qty = int(self.qty.get() or 1)

        con = db(); cur = con.cursor()
        cur.execute("SELECT name, category, mrp, quantity FROM products WHERE product_id=?", (pid,))
        prod = cur.fetchone(); con.close()
        if not prod:
            return

        # check if product already in cart
        existing_item = next((item for item in self.cart if item["pid"] == pid), None)

        if existing_item:
            new_qty = existing_item["qty"] + qty
            if new_qty > prod["quantity"]:
                messagebox.showerror("Stock", f"Only {prod['quantity']} units available in stock.")
                return
            existing_item["qty"] = new_qty
            line_total = new_qty * float(prod["mrp"])

            d_type = existing_item["discount_type"]
            d_val = existing_item["discount_value"]
            discount_amt = d_val if d_type == "Flat" else (line_total * d_val / 100)
            existing_item["final_total"] = max(line_total - discount_amt, 0.0)

            # update Treeview row
            for row in self.cart_tv.get_children():
                row_vals = self.cart_tv.item(row, "values")
                if row_vals[0] == pid:  # update matching row
                    self.cart_tv.item(row, values=(
                        pid, prod["name"], prod["category"], new_qty,
                        f"{prod['mrp']:.2f}", d_type, f"{d_val}", f"{existing_item['final_total']:.2f}"
                    ))
                    break
        else:
            if qty > prod["quantity"]:
                messagebox.showerror("Stock", f"Only {prod['quantity']} units available in stock.")
                return

            mrp = float(prod["mrp"])
            line_total = qty * mrp
            d_type = self.prod_discount_type.get()
            try:
                d_val = float(self.prod_discount_value.get() or 0)
            except:
                d_val = 0.0
            discount_amt = d_val if d_type == "Flat" else (line_total * d_val / 100)
            final_total = max(line_total - discount_amt, 0.0)

            self.cart.append({
                "pid": pid, "name": prod["name"], "cat": prod["category"], "qty": qty,
                "mrp": mrp, "discount_type": d_type, "discount_value": d_val, "final_total": final_total
            })
            self.cart_tv.insert("", "end", values=(
                pid, prod["name"], prod["category"], qty,
                f"{mrp:.2f}", d_type, f"{d_val}", f"{final_total:.2f}"
            ))

        self.update_totals()


    def remove_selected_from_cart(self):
        selected = self.cart_tv.selection()
        if not selected:
            return

        for sel in selected:
            values = self.cart_tv.item(sel, "values")
            pid = values[0]

            # find cart item
            item = next((i for i in self.cart if i["pid"] == pid), None)
            if not item:
                continue

            if item["qty"] > 1:
                item["qty"] -= 1
                line_total = item["qty"] * item["mrp"]
                discount_amt = item["discount_value"] if item["discount_type"] == "Flat" else (line_total * item["discount_value"] / 100)
                item["final_total"] = max(line_total - discount_amt, 0.0)

                self.cart_tv.item(sel, values=(
                    pid, item["name"], item["cat"], item["qty"],
                    f"{item['mrp']:.2f}", item["discount_type"], f"{item['discount_value']}", f"{item['final_total']:.2f}"
                ))
            else:
                # remove completely if qty = 1
                self.cart.remove(item)
                self.cart_tv.delete(sel)

        self.update_totals()


    def remove_selected_from_cart(self):
        sel = self.cart_tv.selection()
        if not sel: return
        index = self.cart_tv.index(sel[0])
        self.cart_tv.delete(sel[0]); del self.cart[index]; self.update_totals()

    def clear_cart(self):
        self.cart = []; self.cart_tv.delete(*self.cart_tv.get_children()); self.update_totals()

    def update_totals(self):
        total = sum(item["final_total"] for item in self.cart)
        self.grand_total_var.set(f"₹{total:.2f}")



    # ---------------- Invoice Generator ----------------
    def generate_invoice_pdf(self, filename, invoice_no, invoice_date,
                             customer_name, customer_phone, items,
                             subtotal, grand_total):
        def safe_float(val):
            try:
                return float(val)
            except Exception:
                return 0.0

        doc = SimpleDocTemplate(filename, pagesize=A4,
                                rightMargin=30, leftMargin=30,
                                topMargin=30, bottomMargin=20)
        story = []
        styles = getSampleStyleSheet()

        # Header
        story.append(Paragraph("<b>LALBAGH ENTERPRISE</b>", styles["Title"]))
        story.append(Paragraph("77, OMRAHGANG, LALBAGH, MURSHIDABAD, WEST BENGAL "
                               "Email: lalbaghenterprise@gmail.com | Phone:78945612300", styles["Normal"]))
        story.append(Spacer(1, 12))

        # Invoice meta
        story.append(Paragraph(f"<b>Invoice No:</b> {invoice_no}", styles["Normal"]))
        story.append(Paragraph(f"<b>Date:</b> {invoice_date}", styles["Normal"]))
        story.append(Paragraph(f"<b>Customer:</b> {customer_name}", styles["Normal"]))
        story.append(Paragraph(f"<b>Phone:</b> {customer_phone}", styles["Normal"]))
        story.append(Spacer(1, 12))

        # Items table
        data = [["Product", "Category", "Qty", "MRP", "Discount", "Final Total (₹)"]]
        for n, c, q, m, d_type, d_val, f in items:
            discount_txt = f"{d_type} {d_val}" if d_val else "-"
            data.append([n, c, q, f"{safe_float(m):.2f}", discount_txt, f"{safe_float(f):.2f}"])

        table = Table(data, colWidths=[150, 100, 50, 70, 100, 100])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightblue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(table)
        story.append(Spacer(1, 12))

        # Totals
        totals_data = [["Subtotal", f"₹ {subtotal:.2f}"],
                       ["Grand Total", f"₹ {grand_total:.2f}"]]
        totals_table = Table(totals_data, colWidths=[300, 200])
        totals_table.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
            ("FONTNAME", (-1, -1), (-1, -1), "Helvetica-Bold"),
            ("TEXTCOLOR", (-1, -1), (-1, -1), colors.green),
            ("FONTSIZE", (-1, -1), (-1, -1), 14),
        ]))
        story.append(totals_table)

        # QR Code
        qr_text = (
            f"Invoice No: {invoice_no}\n"
            f"Date: {invoice_date}\n"
            f"Customer: {customer_name}\n"
            f"Phone: {customer_phone}\n"
            "\n--- Items ---\n" +
            "\n".join([f"{n} ({c}) x{q} @₹{safe_float(m):.2f} - {d_type} {d_val} = ₹{safe_float(f):.2f}"
                       for n, c, q, m, d_type, d_val, f in items]) +
            f"\n\nSubtotal: ₹{subtotal:.2f}\nGrand Total: ₹{grand_total:.2f}"
        )

        qr_code = qr.QrCodeWidget(qr_text)
        bounds = qr_code.getBounds()
        size = 120
        w, h = bounds[2] - bounds[0], bounds[3] - bounds[1]
        d = Drawing(size, size, transform=[size / w, 0, 0, size / h, 0, 0])
        d.add(qr_code)


        story.append(Spacer(1, 16))
        story.append(d)

        doc.build(story)

    # ---------------- Checkout ----------------
    def checkout(self):
        if not self.cart:
            messagebox.showwarning("Checkout", "Cart is empty."); return

        cust_id = None; customer_name = "Walk-in AJ"; customer_phone = ""; customer_email = ""; customer_addr = ""

        if self.has_new_customer_input():
            customer_name = self.new_customer_name.get().strip() or "Walk-in AJ"
            customer_phone = self.new_customer_phone.get().strip()
            customer_email = self.new_customer_email.get().strip()
            customer_addr  = self.new_customer_address.get().strip()

            con = db(); cur = con.cursor()
            new_id = padded_id("customers", "customer_id")
            try:
                cur.execute("""INSERT INTO customers(customer_id,name,phone,email,address) VALUES (?,?,?,?,?)""",
                            (new_id, customer_name, customer_phone, customer_email, customer_addr))
                con.commit(); cust_id = new_id
            except sqlite3.IntegrityError:
                cur.execute("SELECT customer_id,name,phone,email,address FROM customers WHERE phone=? OR email=?",
                            (customer_phone, customer_email))
                r = cur.fetchone()
                if r: cust_id, customer_name, customer_phone, customer_email, customer_addr = \
                    r["customer_id"], r["name"], r["phone"], r["email"], r["address"]
            con.close()

        elif self.customer_sel.get():
            cust_id = self.customer_sel.get().split(" - ", 1)[0]
            con = db(); cur = con.cursor()
            cur.execute("SELECT name,phone,email,address FROM customers WHERE customer_id=?", (cust_id,))
            r = cur.fetchone(); con.close()
            if r: customer_name, customer_phone, customer_email, customer_addr = \
                r["name"], r["phone"], r["email"], r["address"]

        invoice_no = f"INV{int(dt.datetime.now().timestamp())}"; invoice_date = today_str()
        subtotal = sum(item["qty"] * item["mrp"] for item in self.cart)
        grand_total = sum(item["final_total"] for item in self.cart)

        con = db(); cur = con.cursor()
        for item in self.cart:
            cur.execute("""INSERT INTO sales(product_id, product_name, category, quantity, mrp, total_price,
                                             discount_type, discount_value, effective_total, date, sold_by,
                                             customer_name, customer_phone)
                           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (item["pid"], item["name"], item["cat"], item["qty"], item["mrp"],
                         item["qty"] * item["mrp"], item["discount_type"], item["discount_value"],
                         item["final_total"], invoice_date, self.username, customer_name, customer_phone))
            cur.execute("UPDATE products SET quantity = quantity - ? WHERE product_id=?", (item["qty"], item["pid"]))
        con.commit(); con.close()

        filename = filedialog.asksaveasfilename(defaultextension=".pdf", initialfile=f"{invoice_no}.pdf",
                                                filetypes=[("PDF files", "*.pdf")])
        if filename:
            self.generate_invoice_pdf(
                filename, invoice_no, invoice_date, customer_name, customer_phone,
                [(i["name"], i["cat"], i["qty"], i["mrp"], i["discount_type"], i["discount_value"], i["final_total"]) for i in self.cart],
                subtotal, grand_total
            )
            messagebox.showinfo("Invoice", f"Invoice saved:\n{filename}")

        self.clear_cart(); self.refresh(); self.load_customers()
        for v in (self.new_customer_name, self.new_customer_phone, self.new_customer_email, self.new_customer_address):
            v.set(""); self.customer_cmb.set(""); self.customer_sel.set("")

    # ---------------- Refresh History ----------------
    def refresh(self):
        con = db(); cur = con.cursor()
        cur.execute("""SELECT sale_id, date, product_name, category, quantity, mrp, discount_type,
                              discount_value, effective_total, sold_by, customer_name
                       FROM sales ORDER BY sale_id DESC LIMIT 20""")
        rows = [(r["sale_id"], r["date"], r["product_name"], r["category"], r["quantity"], r["mrp"],
                 r["discount_type"], r["discount_value"], r["effective_total"], r["sold_by"], r["customer_name"])
                for r in cur.fetchall()]; con.close()
        insert_rows_striped(self.tv, rows)

    def show_returns(self):
        win = tk.Toplevel(self)
        win.title("Product Return / Refund")
        win.geometry("750x520")

        sale_id = tk.StringVar()
        refund_data = []
        reason = tk.StringVar()

        # --- Sale details tree ---
        cols = ("Product", "Qty Sold", "Unit Price", "Refund Qty", "Refund Amount")
        tree = ttk.Treeview(win, columns=cols, show="headings")
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, anchor="center", width=120)
        tree.pack(fill="both", expand=True, pady=10)

        # ✅ Load sale details
        def load_sale():
            for row in tree.get_children():
                tree.delete(row)
            refund_data.clear()

            con = db()
            cur = con.cursor()
            cur.execute("""
                        SELECT sale_id, product_name, quantity, effective_total, mrp, product_id
                        FROM sales
                        WHERE sale_id = ?
                        """, (sale_id.get(),))
            rows = cur.fetchall()
            con.close()

            if not rows:
                messagebox.showerror("Error", "Sale not found")
                return

            for sid, name, qty, total, mrp, pid in rows:
                refund_data.append([sid, pid, name, qty, mrp, 0, 0.0])  # refund qty, refund amt
                tree.insert("", "end", values=(name, qty, f"₹{mrp:.2f}", 0, "₹0.00"))

            # Allow editing refund qty
            def set_refund(event):
                item = tree.selection()
                if not item:
                    return
                idx = tree.index(item[0])
                old = refund_data[idx]
                qty_win = tk.simpledialog.askinteger(
                    "Refund Quantity",
                    f"Enter refund qty for {old[2]} (max {old[3]}):",
                    minvalue=1, maxvalue=old[3]
                )
                if qty_win:
                    refund_amt = qty_win * old[4]
                    refund_data[idx][5] = qty_win
                    refund_data[idx][6] = refund_amt
                    tree.item(item[0], values=(old[2], old[3], f"₹{old[4]:.2f}", qty_win, f"₹{refund_amt:.2f}"))

            tree.bind("<Double-1>", set_refund)

        # --- Recent sales dropdown ---
        f_top = tk.Frame(win, bg=THEME["bg"])
        f_top.pack(fill="x", pady=6)

        tk.Label(f_top, text="Select Recent Sale:", bg=THEME["bg"], font=FONT_MD).pack(side="left", padx=4)
        sale_combo = ttk.Combobox(f_top, textvariable=sale_id, width=40, state="readonly")
        sale_combo.pack(side="left", padx=5)

        con = db()
        cur = con.cursor()
        cur.execute("""
                    SELECT sale_id, product_name, date, effective_total, customer_name
                    FROM sales
                    ORDER BY sale_id DESC
                        LIMIT 50
                    """)
        rows = cur.fetchall()
        con.close()

        if rows:
            formatted = [
                f"{r[0]} - {r[1]} - {r[2]} - ₹{r[3]:.2f} - {r[4] if r[4] else ''}"
                for r in rows
            ]
            ids = [str(r[0]) for r in rows]
            sale_combo["values"] = formatted

            def set_sale(event):
                idx = sale_combo.current()
                if idx >= 0:
                    sale_id.set(ids[idx])
                    load_sale()

            sale_combo.bind("<<ComboboxSelected>>", set_sale)

        # --- Reason for return ---
        f_reason = tk.Frame(win, bg=THEME["bg"])
        f_reason.pack(fill="x", pady=6)
        tk.Label(f_reason, text="Reason for Return:", bg=THEME["bg"], font=FONT_MD).pack(side="left", padx=4)
        tk.Entry(f_reason, textvariable=reason, width=40).pack(side="left", padx=5)

        # ✅ Confirm Refund
        def process_refund():
            if not refund_data:
                messagebox.showerror("Error", "No sale loaded")
                return
            any_refund = False

            con = db()
            cur = con.cursor()
            today = dt.datetime.now().strftime("%Y-%m-%d")

            for sid, pid, name, qty, mrp, r_qty, r_amt in refund_data:
                if r_qty > 0:
                    any_refund = True
                    # Save in returns table
                    cur.execute("""
                                INSERT INTO returns (sale_id, product_id, quantity, refund_amount, date, reason)
                                VALUES (?, ?, ?, ?, ?, ?)
                                """, (sid, pid, r_qty, r_amt, today, reason.get()))
                    # Update stock back
                    cur.execute("UPDATE products SET quantity = quantity + ? WHERE product_id = ?", (r_qty, pid))

            con.commit()
            con.close()

            if any_refund:
                messagebox.showinfo("Success", "Refund processed successfully")
                win.destroy()
            else:
                messagebox.showwarning("No Action", "No refund quantities entered")

        # --- Confirm Button ---
        tk.Button(win, text="Confirm Refund", bg=THEME["danger"], fg="white",
                  font=FONT_MD, command=process_refund).pack(pady=10)

    def show_returns(self):
        """Open a refund window, search sales, and process returns."""
        win = tk.Toplevel(self)
        win.title("Process Refund / Return")
        win.geometry("700x400")
        win.configure(bg=THEME["bg"])

        # --- Sale search ---
        tk.Label(win, text="Enter Sale ID:", font=FONT_MD, bg=THEME["bg"]).pack(pady=6)
        sale_id_var = tk.StringVar()
        tk.Entry(win, textvariable=sale_id_var, font=FONT_MD, width=15).pack()

        cols = ("product_id", "product_name", "quantity", "effective_total")
        tv = ttk.Treeview(win, columns=cols, show="headings", height=8)
        for c, w in zip(cols, [100, 200, 100, 120]):
            tv.heading(c, text=c.title())
            tv.column(c, width=w, anchor="center")
        tv.pack(fill="x", padx=10, pady=10)
        setup_treeview_striped(tv)

        def load_sale():
            sid = sale_id_var.get().strip()
            if not sid:
                return
            con = db();
            cur = con.cursor()
            cur.execute("""SELECT sale_id, product_id, product_name, quantity, effective_total
                           FROM sales
                           WHERE sale_id = ?""", (sid,))
            rows = cur.fetchall()
            con.close()
            tv.delete(*tv.get_children())
            for r in rows:
                tv.insert("", "end", values=(r["product_id"], r["product_name"], r["quantity"], r["effective_total"]))

        tk.Button(win, text="Load Sale", font=FONT_MD,
                  bg=THEME["primary"], fg="white", command=load_sale).pack(pady=6)

        # --- Refund Form ---
        form = tk.Frame(win, bg=THEME["bg"])
        form.pack(fill="x", padx=10, pady=10)

        refund_qty_var = tk.StringVar()
        reason_var = tk.StringVar()

        tk.Label(form, text="Refund Quantity:", font=FONT_MD, bg=THEME["bg"]).grid(row=0, column=0, padx=6, pady=6)
        tk.Entry(form, textvariable=refund_qty_var, font=FONT_MD, width=8).grid(row=0, column=1, padx=6, pady=6)

        tk.Label(form, text="Reason:", font=FONT_MD, bg=THEME["bg"]).grid(row=0, column=2, padx=6, pady=6)
        tk.Entry(form, textvariable=reason_var, font=FONT_MD, width=40).grid(row=0, column=3, padx=6, pady=6)

        def do_refund():
            sel = tv.selection()
            if not sel:
                messagebox.showwarning("Refund", "Select a product to refund.")
                return
            vals = tv.item(sel[0], "values")
            pid = vals[0]
            sid = sale_id_var.get().strip()
            try:
                qty = int(refund_qty_var.get().strip())
            except:
                messagebox.showerror("Refund", "Invalid quantity.")
                return
            reason = reason_var.get().strip()
            self.process_refund(sid, pid, qty, reason)
            load_sale()  # refresh table after refund

        tk.Button(win, text="Process Refund", font=FONT_MD,
                  bg=THEME["warning"], fg="black", command=do_refund).pack(pady=10)

    def process_refund(self, sale_id, product_id, refund_qty, reason=""):
        """Validate and process a refund securely."""
        if not reason.strip():
            messagebox.showwarning("Refund", "Refund reason is required.")
            return

        con = db();
        cur = con.cursor()
        try:
            # --- Fetch sale record ---
            cur.execute("""SELECT quantity, effective_total
                           FROM sales
                           WHERE sale_id = ?
                             AND product_id = ?""", (sale_id, product_id))
            row = cur.fetchone()
            if not row:
                messagebox.showerror("Refund", "Sale not found.")
                return

            sold_qty, eff_total = row["quantity"], row["effective_total"]

            # --- Already refunded qty ---
            cur.execute("""SELECT IFNULL(SUM(quantity), 0) AS refunded
                           FROM returns
                           WHERE sale_id = ?
                             AND product_id = ?""", (sale_id, product_id))
            already_refunded = cur.fetchone()["refunded"]

            # --- Validate qty ---
            if refund_qty <= 0:
                messagebox.showerror("Refund", "Refund qty must be > 0.")
                return
            if refund_qty + already_refunded > sold_qty:
                messagebox.showerror(
                    "Refund",
                    f"Cannot refund {refund_qty}. Already refunded {already_refunded} of {sold_qty}."
                )
                return

            # --- Refund amount per unit ---
            unit_price = eff_total / sold_qty
            refund_amt = round(unit_price * refund_qty, 2)

            # --- Transaction ---
            cur.execute("""INSERT INTO returns
                               (sale_id, product_id, quantity, refund_amount, date, reason)
                           VALUES (?, ?, ?, ?, ?, ?)""",
                        (sale_id, product_id, refund_qty, refund_amt, today_str(), reason))
            cur.execute("UPDATE products SET quantity = quantity + ? WHERE product_id=?",
                        (refund_qty, product_id))
            con.commit()

            messagebox.showinfo("Refund", f"Refund processed: ₹{refund_amt:.2f}")
        except Exception as e:
            con.rollback()
            messagebox.showerror("Refund", f"Error: {e}")
        finally:
            con.close()

# ---------- REPORTS DASHBOARD ----------
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import datetime as dt
import openpyxl
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import matplotlib.pyplot as plt
import os
from tkcalendar import DateEntry
import tkinter as tk
from tkinter import messagebox, filedialog
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import datetime as dt


# ---------- HELPERS ----------
def export_treeview_to_excel(tree, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    # headers
    ws.append([tree.heading(col)["text"] for col in tree["columns"]])
    # rows
    for row_id in tree.get_children():
        ws.append(tree.item(row_id)["values"])
    # footer timestamp
    ws.append([])
    ws.append([f"Exported On: {dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    wb.save(filename)


def export_treeview_to_pdf(tree, filename, title="Report"):
    doc = SimpleDocTemplate(filename, pagesize=A4)
    data = []
    data.append([tree.heading(col)["text"] for col in tree["columns"]])
    for row_id in tree.get_children():
        data.append(tree.item(row_id)["values"])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ]))

    timestamp = f"Exported On: {dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    story = [Paragraph(title, getSampleStyleSheet()["Title"]), Spacer(1, 12),
             table, Spacer(1, 12), Paragraph(timestamp, getSampleStyleSheet()["Normal"])]
    doc.build(story)


# ---------- REPORTS DASHBOARD ----------
class SectionReports(tk.Frame):
    def __init__(self, parent, user, role):
        if role != "Admin":
            messagebox.showerror("Access Denied", "Reports are available only for Admin users.")
            parent.destroy()
            return

        super().__init__(parent, bg=THEME["bg"])
        self.pack(fill="both", expand=True)
        self.user = user

        # --- TITLE ---
        title = tk.Label(self, text="Reports Dashboard (Admin Only)",
                         font=FONT_XL, bg=THEME["bg"], fg=THEME["dark"])
        title.pack(pady=10)

        # --- KPI SUMMARY ---
        kpi_frame = tk.Frame(self, bg=THEME["bg"])
        kpi_frame.pack(fill="x", pady=10)

        self.kpi_sales = tk.Label(kpi_frame, text="Total Sales: ₹0.00",
                                  font=FONT_LG, bg=THEME["primary"], fg="white", width=30, relief="groove")
        self.kpi_sales.grid(row=0, column=0, padx=8, pady=8)

        self.kpi_customers = tk.Label(kpi_frame, text="Total Customers: 0",
                                      font=FONT_LG, bg=THEME["success"], fg="white", width=30, relief="groove")
        self.kpi_customers.grid(row=0, column=1, padx=8, pady=8)

        self.kpi_profit = tk.Label(kpi_frame, text="Profit Margin: ₹0.00",
                                   font=FONT_LG, bg=THEME["danger"], fg="white", width=30, relief="groove")
        self.kpi_profit.grid(row=0, column=2, padx=8, pady=8)

        # --- CHART BUTTONS ---
        chart_frame = tk.Frame(self, bg=THEME["bg"])
        chart_frame.pack(fill="both", expand=True, padx=12, pady=12)

        tk.Button(chart_frame, text="Monthly Sales Trend",
                  font=FONT_MD, bg=THEME["primary"], fg="white",
                  command=self.show_sales_trend).grid(row=0, column=0, padx=10, pady=10)

        tk.Button(chart_frame, text="Top 5 Products",
                  font=FONT_MD, bg=THEME["success"], fg="white",
                  command=self.show_top_products).grid(row=0, column=1, padx=10, pady=10)

        tk.Button(chart_frame, text="Supplier vs Supplier",
                  font=FONT_MD, bg=THEME["warning"], fg="black",
                  command=self.show_supplier_comparison).grid(row=0, column=2, padx=10, pady=10)

        tk.Button(chart_frame, text="Profit Margin Report",
                  font=FONT_MD, bg=THEME["danger"], fg="white",
                  command=self.show_profit_report).grid(row=0, column=3, padx=10, pady=10)
        tk.Button(chart_frame, text="Product Sales Share",
                  font=FONT_MD, bg=THEME["accent"], fg="white",
                  command=self.show_product_sales_share).grid(row=0, column=4, padx=10, pady=10)
        tk.Button(chart_frame, text="Daily Sales Trend",
                  font=FONT_MD, bg=THEME["primary"], fg="white",
                  command=self.show_daily_sales_trend).grid(row=0, column=5, padx=10, pady=10)
        tk.Button(chart_frame, text="Export All Reports",
                  font=FONT_MD, bg=THEME["success"], fg="white",
                  command=self.export_all_reports).grid(row=0, column=6, padx=10, pady=10)
        tk.Button(chart_frame, text="Profit Analysis Report",
                  font=FONT_MD, bg="#8E44AD", fg="white",
                  command=self.show_profit_analysis).grid(row=0, column=8, padx=10, pady=10)

        # --- SALES HISTORY ---
        history_frame = tk.LabelFrame(self, text="Sales History", bg=THEME["bg"], font=FONT_LG, fg=THEME["dark"])
        history_frame.pack(fill="both", expand=True, padx=12, pady=12)

        filter_frame = tk.Frame(history_frame, bg=THEME["bg"])
        filter_frame.pack(fill="x", padx=8, pady=4)

        self.f_from = tk.StringVar(value=today_str())
        self.f_to = tk.StringVar(value=today_str())

        tk.Label(filter_frame, text="From:", bg=THEME["bg"], font=FONT_MD).pack(side="left", padx=4)
        tk.Entry(filter_frame, textvariable=self.f_from, width=12, font=FONT_MD).pack(side="left")
        tk.Label(filter_frame, text="To:", bg=THEME["bg"], font=FONT_MD).pack(side="left", padx=4)
        tk.Entry(filter_frame, textvariable=self.f_to, width=12, font=FONT_MD).pack(side="left")

        tk.Button(filter_frame, text="Apply", font=FONT_MD, bg=THEME["primary"], fg="white",
                  command=self.refresh_sales).pack(side="left", padx=6)

        btn_frame = tk.Frame(history_frame, bg=THEME["bg"])
        btn_frame.pack(fill="x", padx=8, pady=4, anchor="e")

        tk.Button(btn_frame, text="Export Excel", font=FONT_MD,
                  command=self.export_sales_excel, bg=THEME["accent"], fg="white").pack(side="right", padx=6)
        tk.Button(btn_frame, text="Export PDF", font=FONT_MD,
                  command=self.export_sales_pdf, bg=THEME["primary"], fg="white").pack(side="right", padx=6)

        cols = ("sale_id", "date", "product_name", "category", "quantity", "mrp",
                "effective_total", "sold_by", "customer_name", "customer_phone")
        self.sales_tv = ttk.Treeview(history_frame, columns=cols, show="headings", height=12)
        for c, w in zip(cols, [60, 90, 200, 120, 70, 70, 100, 100, 160, 120]):
            self.sales_tv.heading(c, text=c.replace("_", " ").title())
            self.sales_tv.column(c, width=w, anchor="center")
        self.sales_tv.pack(fill="both", expand=True, pady=8)
        setup_treeview_striped(self.sales_tv)

        # Load data
        self.refresh_summary()
        self.refresh_sales()

    # --- KPI REFRESH ---
    def refresh_summary(self):
        con = db();
        cur = con.cursor()
        cur.execute("SELECT IFNULL(SUM(effective_total),0) FROM sales WHERE strftime('%m',date)=strftime('%m','now')")
        total_sales = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM customers")
        total_customers = cur.fetchone()[0]

        cur.execute("""SELECT IFNULL(SUM(s.effective_total - (p.unit_price * s.quantity)), 0)
                       FROM sales s
                                JOIN products p ON s.product_id = p.product_id""")
        profit = cur.fetchone()[0]
        con.close()
        self.kpi_sales.config(text=f"Total Sales: ₹{total_sales:.2f}")
        self.kpi_customers.config(text=f"Total Customers: {total_customers}")
        self.kpi_profit.config(text=f"Profit Margin: ₹{profit:.2f}")

    # --- CHARTS EMBEDDED IN TKINTER ---
    # --- CHARTS EMBEDDED IN TKINTER (WITH EXPORT) ---
    import tkinter as tk
    from tkinter import filedialog, messagebox
    from matplotlib.figure import Figure
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    import datetime as dt
    import locale
    from matplotlib.ticker import FuncFormatter
    from matplotlib.ticker import FuncFormatter

    # Set Indian numbering locale (safe fallback)
    try:
        locale.setlocale(locale.LC_ALL, 'en_IN.UTF-8')
    except:
        locale.setlocale(locale.LC_ALL, '')

    def show_sales_trend(self):
        con = db()
        cur = con.cursor()
        cur.execute("""
            SELECT strftime('%Y-%m', date) as month, SUM(effective_total)
            FROM sales
            GROUP BY month
            ORDER BY month
        """)
        rows = cur.fetchall()
        con.close()

        if not rows:
            messagebox.showwarning("No Data", "No sales data available to display.")
            return

        months = [r[0] for r in rows]
        totals = [r[1] for r in rows]

        win = tk.Toplevel(self)
        win.title("Monthly Sales Trend")
        win.geometry("750x550")

        fig = Figure(figsize=(6.5, 4.5), dpi=100)
        ax = fig.add_subplot(111)

        # --- Smart Currency Formatter for ₹, Lakhs, Crores ---
        def rupee_smart_format(x, pos=None):
            if x >= 1_00_00_000:
                return f"₹{x / 1_00_00_000:.1f} Cr"
            elif x >= 1_00_000:
                return f"₹{x / 1_00_000:.1f} L"
            else:
                return f"₹{int(x):,}"

        ax.yaxis.set_major_formatter(FuncFormatter(rupee_smart_format))

        # --- Draw colored trend lines ---
        for i in range(1, len(totals)):
            color = "green" if totals[i] >= totals[i - 1] else "red"
            ax.plot(months[i - 1:i + 1], totals[i - 1:i + 1],
                    marker="o", color=color, linewidth=2)

        # Add title and labels
        ax.set_title("📈 Monthly Sales Trend (₹)", fontsize=12, fontweight='bold')
        ax.set_xlabel("Month", fontsize=10)
        ax.set_ylabel("Sales Amount (₹)", fontsize=10)
        ax.tick_params(axis="x", rotation=45)

        # Add export date (top-right)
        export_date = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ax.text(
            0.99, 0.99, f"Exported On: {export_date}",
            ha="right", va="top", transform=ax.transAxes,
            fontsize=8, color="gray"
        )

        # Add value labels above each point
        for i, total in enumerate(totals):
            label = rupee_smart_format(total)
            ax.text(i, total, label, ha='center', va='bottom', fontsize=8,
                    color='darkgreen' if i == 0 or (i > 0 and totals[i] >= totals[i - 1]) else 'darkred')

        # Add subtle background grid
        ax.grid(True, linestyle='--', alpha=0.5)

        # Draw chart in Tkinter
        canvas = FigureCanvasTkAgg(fig, win)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)

        # --- Export Functionality ---
        def export_chart(fmt):
            save_path = filedialog.asksaveasfilename(
                defaultextension=f".{fmt}",
                filetypes=[(fmt.upper(), f"*.{fmt}")],
                initialfile=f"monthly_sales_trend.{fmt}"
            )
            if save_path:
                fig.savefig(save_path, format=fmt)
                messagebox.showinfo("Export", f"Chart exported as {save_path}")

        # --- Export Buttons ---
        btn_frame = tk.Frame(win, bg=THEME["bg"])
        btn_frame.pack(pady=6)

        tk.Button(btn_frame, text="Export PNG", command=lambda: export_chart("png")).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Export JPEG", command=lambda: export_chart("jpeg")).pack(side="left", padx=5)

    from tkcalendar import DateEntry
    import tkinter as tk
    from tkinter import messagebox, filedialog
    from matplotlib.figure import Figure
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    import datetime as dt

    def show_top_products(self):
        win = tk.Toplevel(self)
        win.title("Top 5 Products")
        win.geometry("750x550")

        # --- Filter bar with tkcalendar ---
        f_top = tk.Frame(win, bg=THEME["bg"])
        f_top.pack(fill="x", pady=6)

        tk.Label(f_top, text="From:", bg=THEME["bg"], font=FONT_MD).pack(side="left", padx=4)
        from_date = DateEntry(f_top, width=12, date_pattern="yyyy-mm-dd")
        from_date.set_date(dt.date(2000, 1, 1))  # default earliest date
        from_date.pack(side="left")

        tk.Label(f_top, text="To:", bg=THEME["bg"], font=FONT_MD).pack(side="left", padx=4)
        to_date = DateEntry(f_top, width=12, date_pattern="yyyy-mm-dd")
        to_date.set_date(dt.date.today())
        to_date.pack(side="left")

        # --- Matplotlib figure ---
        fig = Figure(figsize=(6.5, 4.5), dpi=100)
        ax1 = fig.add_subplot(111)

        def load_chart():
            con = db()
            cur = con.cursor()
            cur.execute("""
                SELECT product_name, SUM(quantity) AS qty
                FROM sales
                WHERE date BETWEEN ? AND ?
                GROUP BY product_id, product_name
                ORDER BY qty DESC
                LIMIT 5
            """, (from_date.get(), to_date.get()))
            rows = cur.fetchall()
            con.close()

            if not rows:
                messagebox.showinfo("No Data", "No sales found for this date range.")
                ax1.clear()
                canvas.draw()
                return

            names = [r[0] for r in rows]
            qtys = [r[1] for r in rows]

            ax1.clear()
            ax1.bar(names, qtys, color="skyblue")
            ax1.set_title("Top 5 Products by Quantity Sold")
            ax1.set_xlabel("Product")
            ax1.set_ylabel("Quantity")
            ax1.tick_params(axis="x", rotation=30)

            # Exported date top-right
            export_date = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ax1.text(0.99, 0.99, f"Exported On: {export_date}",
                     ha="right", va="top", transform=ax1.transAxes,
                     fontsize=8, color="gray")

            canvas.draw()

        # --- Canvas ---
        canvas = FigureCanvasTkAgg(fig, win)
        canvas.get_tk_widget().pack(fill="both", expand=True)
        canvas.draw()

        # --- Buttons ---
        btn_frame = tk.Frame(win, bg=THEME["bg"])
        btn_frame.pack(pady=6)
        tk.Button(btn_frame, text="Apply Filter", command=load_chart).pack(side="left", padx=5)

        def export_chart(fmt):
            save_path = filedialog.asksaveasfilename(
                defaultextension=f".{fmt}",
                filetypes=[(fmt.upper(), f"*.{fmt}")],
                initialfile=f"top_products.{fmt}"
            )
            if save_path:
                fig.savefig(save_path, format=fmt)
                messagebox.showinfo("Export", f"Chart exported as {save_path}")

        tk.Button(btn_frame, text="Export PNG", command=lambda: export_chart("png")).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Export JPEG", command=lambda: export_chart("jpeg")).pack(side="left", padx=5)

        # --- Load first view ---
        load_chart()

    def show_supplier_comparison(self):
        win = tk.Toplevel(self)
        win.title("Supplier vs Supplier")
        win.geometry("750x550")

        f_top = tk.Frame(win, bg=THEME["bg"])
        f_top.pack(fill="x", pady=6)
        from_date = tk.StringVar(value="2000-01-01")
        to_date = tk.StringVar(value=today_str())
        tk.Label(f_top, text="From:", bg=THEME["bg"], font=FONT_MD).pack(side="left", padx=4)
        tk.Entry(f_top, textvariable=from_date, width=12).pack(side="left")
        tk.Label(f_top, text="To:", bg=THEME["bg"], font=FONT_MD).pack(side="left", padx=4)
        tk.Entry(f_top, textvariable=to_date, width=12).pack(side="left")

        fig = Figure(figsize=(6.5, 4.5), dpi=100)
        ax = fig.add_subplot(111)

        def load_chart():
            con = db()
            cur = con.cursor()
            cur.execute("""SELECT s.company, SUM(sl.effective_total) as total_sales
                           FROM sales sl
                                    JOIN products p ON sl.product_id = p.product_id
                                    JOIN suppliers s ON p.supplier_id = s.supplier_id
                           WHERE sl.date BETWEEN ? AND ?
                           GROUP BY s.supplier_id
                           ORDER BY total_sales DESC""", (from_date.get(), to_date.get()))
            rows = cur.fetchall()
            con.close()

            suppliers = [r[0] for r in rows]
            totals = [r[1] for r in rows]

            ax.clear()
            ax.bar(suppliers, totals, color="orange")
            ax.set_title("Supplier vs Supplier Sales")
            ax.set_xlabel("Supplier")
            ax.set_ylabel("Sales ₹")
            ax.tick_params(axis="x", rotation=30)

            # Exported date top-right
            export_date = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ax.text(0.99, 0.99, f"Exported On: {export_date}",
                    ha="right", va="top", transform=ax.transAxes, fontsize=8, color="gray")

            canvas.draw()

        canvas = FigureCanvasTkAgg(fig, win)
        canvas.get_tk_widget().pack(fill="both", expand=True)
        canvas.draw()

        btn_frame = tk.Frame(win, bg=THEME["bg"])
        btn_frame.pack(pady=6)
        tk.Button(btn_frame, text="Apply Filter", command=load_chart).pack(side="left", padx=5)

        def export_chart(fmt):
            save_path = filedialog.asksaveasfilename(
                defaultextension=f".{fmt}",
                filetypes=[(fmt.upper(), f"*.{fmt}")],
                initialfile=f"supplier_comparison.{fmt}"
            )
            if save_path:
                fig.savefig(save_path, format=fmt)
                messagebox.showinfo("Export", f"Chart exported as {save_path}")

        tk.Button(btn_frame, text="Export PNG", command=lambda: export_chart("png")).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Export JPEG", command=lambda: export_chart("jpeg")).pack(side="left", padx=5)

        # Load first view
        load_chart()

    # --- PROFIT REPORT ---
    def show_profit_report(self):
        win = tk.Toplevel(self);
        win.title("Profit Margin Report")
        win.geometry("1050x600");
        win.configure(bg=THEME["bg"])

        f_top = tk.Frame(win, bg=THEME["bg"]);
        f_top.pack(fill="x", pady=6)
        from_date = tk.StringVar(value=today_str())
        to_date = tk.StringVar(value=today_str())
        tk.Label(f_top, text="From:", bg=THEME["bg"], font=FONT_MD).pack(side="left", padx=4)
        tk.Entry(f_top, textvariable=from_date, width=12).pack(side="left")
        tk.Label(f_top, text="To:", bg=THEME["bg"], font=FONT_MD).pack(side="left", padx=4)
        tk.Entry(f_top, textvariable=to_date, width=12).pack(side="left")

        cols = ("Date", "Product", "Unit Price", "Quantity", "Grand Total", "Profit", "Profit %")
        tv = ttk.Treeview(win, columns=cols, show="headings", height=15)
        for c in cols:
            tv.heading(c, text=c);
            tv.column(c, anchor="center", width=120)
        tv.pack(fill="both", expand=True, padx=10, pady=10)
        setup_treeview_striped(tv)

        def load_data():
            f, t = from_date.get().strip(), to_date.get().strip()
            con = db();
            cur = con.cursor()
            cur.execute("""
                        SELECT s.date,
                               s.product_name,
                               p.unit_price,
                               s.quantity,
                               s.effective_total,
                               (s.effective_total - (p.unit_price * s.quantity)) AS profit
                        FROM sales s
                                 JOIN products p ON p.product_id = s.product_id
                        WHERE s.date BETWEEN ? AND ?
                        ORDER BY s.date
                        """, (f, t))
            rows = cur.fetchall();
            con.close()

            tv.delete(*tv.get_children())
            for r in rows:
                grand_total = r["effective_total"];
                profit = r["profit"]
                profit_pct = (profit / grand_total * 100) if grand_total else 0
                tv.insert("", "end", values=(r["date"], r["product_name"],
                                             f"{r['unit_price']:.2f}", r["quantity"],
                                             f"{grand_total:.2f}", f"{profit:.2f}", f"{profit_pct:.2f}%"))

        btn_frame = tk.Frame(win, bg=THEME["bg"]);
        btn_frame.pack(pady=8)
        tk.Button(btn_frame, text="Apply Date Filter", bg=THEME["primary"], fg="white",
                  command=load_data).pack(side="left", padx=6)
        tk.Button(btn_frame, text="Export Profit (Excel)",
                  command=lambda: export_treeview_to_excel(tv, "profit_report.xlsx")).pack(side="left", padx=6)
        tk.Button(btn_frame, text="Export Profit (PDF)",
                  command=lambda: self.export_profit_pdf(tv, from_date.get(), to_date.get())).pack(side="left", padx=6)

        load_data()


    def export_profit_pdf(self, tv, f, t):
        export_date = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        save_path = filedialog.asksaveasfilename(defaultextension=".pdf",
                                                 initialfile="profit_margin.pdf",
                                                 filetypes=[("PDF", "*.pdf")])
        if not save_path: return

        cols = ("Date", "Product", "Unit Price", "Quantity", "Grand Total", "Profit", "Profit %")
        data = [cols]
        total_profit = total_sales = 0.0
        profit_rows = []

        for child in tv.get_children():
            vals = tv.item(child, "values")
            data.append(list(vals))
            try:
                grand_total = float(vals[4]);
                profit = float(vals[5])
                total_sales += grand_total;
                total_profit += profit
                profit_pct = (profit / grand_total * 100) if grand_total else 0
                profit_rows.append((vals[1], profit_pct))
            except:
                pass

        doc = SimpleDocTemplate(save_path, pagesize=A4, rightMargin=24, leftMargin=24,
                                topMargin=24, bottomMargin=24)
        styles = getSampleStyleSheet()
        story = [Paragraph("<b>Profit Margin Report</b>", styles["Title"]), Spacer(1, 8)]
        story.append(Paragraph(f"<b>Date Range:</b> {f} → {t}", styles["Normal"]))
        story.append(Paragraph(f"<b>Exported On:</b> {export_date}", styles["Normal"]))
        story.append(Spacer(1, 12))

        tbl = Table(data, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.black),
            ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        story.append(tbl);
        story.append(Spacer(1, 12))

        profit_pct_total = (total_profit / total_sales * 100) if total_sales else 0
        summary = Table([
            [f"Total Sales ({f} → {t})", f"₹{total_sales:.2f}"],
            [f"Total Profit ({f} → {t})", f"₹{total_profit:.2f}"],
            ["Overall Profit %", f"{profit_pct_total:.2f}%"]
        ], colWidths=[300, 150])
        summary.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgreen),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        story.append(summary);
        story.append(Spacer(1, 20))

        # --- Add Top 5 Products by Profit % chart ---
        if profit_rows:
            top_rows = sorted(profit_rows, key=lambda x: x[1], reverse=True)[:5]
            products = [r[0] for r in top_rows];
            percents = [r[1] for r in top_rows]

            plt.figure(figsize=(5, 3))
            plt.bar(products, percents, color="purple")
            plt.title("Top 5 Products by Profit %");
            plt.xlabel("Product");
            plt.ylabel("Profit %")
            plt.xticks(rotation=30);
            plt.tight_layout()

            chart_path = "temp_profit_chart.png"
            plt.savefig(chart_path);
            plt.close()
            story.append(Paragraph("<b>Top 5 Products by Profit %</b>", styles["Heading3"]))
            story.append(RLImage(chart_path, width=400, height=250));
            story.append(Spacer(1, 12))
            doc.build(story);
            os.remove(chart_path)
        else:
            doc.build(story)

    # --- SALES HISTORY ---
    def refresh_sales(self):
        f1, f2 = self.f_from.get().strip(), self.f_to.get().strip()
        con = db();
        cur = con.cursor()
        cur.execute(
            """SELECT sale_id, date, product_name, category, quantity, mrp, effective_total, sold_by, customer_name, customer_phone
               FROM sales WHERE date BETWEEN ? AND ? ORDER BY sale_id DESC""", (f1, f2))
        rows = [(r["sale_id"], r["date"], r["product_name"], r["category"],
                 r["quantity"], f"{r['mrp']:.2f}", f"{r['effective_total']:.2f}",
                 r["sold_by"], r["customer_name"], r["customer_phone"]) for r in cur.fetchall()]
        con.close()
        insert_rows_striped(self.sales_tv, rows)

    def export_sales_excel(self):
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if save_path:
            export_treeview_to_excel(self.sales_tv, save_path)
            messagebox.showinfo("Export", f"Sales exported to Excel:\n{save_path}")

    def export_sales_pdf(self):
        save_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF Files", "*.pdf")])
        if save_path:
            f1, f2 = self.f_from.get().strip(), self.f_to.get().strip()
            export_treeview_to_pdf(self.sales_tv, save_path, f"Sales Report ({f1} → {f2})")
            messagebox.showinfo("Export", f"Sales exported to PDF:\n{save_path}")

    def show_product_sales_share(self):
        win = tk.Toplevel(self)
        win.title("Product Sales Share (Pie Chart)")
        win.geometry("750x600")

        # --- Date filter frame ---
        f_top = tk.Frame(win, bg=THEME["bg"])
        f_top.pack(fill="x", pady=6)
        from_date = tk.StringVar(value="2000-01-01")
        to_date = tk.StringVar(value=today_str())
        tk.Label(f_top, text="From:", bg=THEME["bg"], font=FONT_MD).pack(side="left", padx=4)
        tk.Entry(f_top, textvariable=from_date, width=12).pack(side="left")
        tk.Label(f_top, text="To:", bg=THEME["bg"], font=FONT_MD).pack(side="left", padx=4)
        tk.Entry(f_top, textvariable=to_date, width=12).pack(side="left")

        fig = Figure(figsize=(6.5, 5.2), dpi=100)
        ax = fig.add_subplot(111)
        fig.subplots_adjust(bottom=0.18)  # leave space at bottom

        def load_chart():
            con = db()
            cur = con.cursor()
            cur.execute("""
                        SELECT product_name, SUM(effective_total) as total_sales
                        FROM sales
                        WHERE date BETWEEN ? AND ?
                        GROUP BY product_id
                        ORDER BY total_sales DESC
                        """, (from_date.get(), to_date.get()))
            rows = cur.fetchall()
            con.close()

            labels = [r[0] for r in rows]
            totals = [r[1] for r in rows]
            grand_total = sum(totals)

            ax.clear()
            fig.suptitle("")  # reset previous total text

            if totals:
                ax.pie(
                    totals, labels=labels, autopct="%1.1f%%", startangle=140
                )
                ax.set_title("Product-wise Sales Share", fontsize=12, fontweight="bold")

                # Show total sales BELOW chart (figure-level, not inside axes)
                fig.suptitle(f"Total Sales: ₹{grand_total:,.2f}",
                             x=0.5, y=0.02, ha="center",
                             fontsize=11, fontweight="bold", color="green")

                # Exported date top-right inside chart
                export_date = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ax.text(0.99, 0.99, f"Exported On: {export_date}",
                        ha="right", va="top", transform=ax.transAxes,
                        fontsize=8, color="gray")
            else:
                ax.text(0.5, 0.5, "No data in selected range",
                        ha="center", va="center", fontsize=12)

            canvas.draw()

        canvas = FigureCanvasTkAgg(fig, win)
        canvas.get_tk_widget().pack(fill="both", expand=True)
        canvas.draw()

        # Buttons
        btn_frame = tk.Frame(win, bg=THEME["bg"])
        btn_frame.pack(pady=6)
        tk.Button(btn_frame, text="Apply Filter", command=load_chart).pack(side="left", padx=5)

        def export_chart(fmt):
            save_path = filedialog.asksaveasfilename(
                defaultextension=f".{fmt}",
                filetypes=[(fmt.upper(), f"*.{fmt}")],
                initialfile=f"product_sales_share.{fmt}"
            )
            if save_path:
                fig.savefig(save_path, format=fmt)
                messagebox.showinfo("Export", f"Chart exported as {save_path}")

        tk.Button(btn_frame, text="Export PNG", command=lambda: export_chart("png")).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Export JPEG", command=lambda: export_chart("jpeg")).pack(side="left", padx=5)

        # Load first view
        load_chart()
    def show_daily_sales_trend(self):
        win = tk.Toplevel(self)
        win.title("Daily Sales Trend")
        win.geometry("800x600")

        # --- Date filter frame ---
        f_top = tk.Frame(win, bg=THEME["bg"])
        f_top.pack(fill="x", pady=6)
        from_date = tk.StringVar(value="2000-01-01")
        to_date = tk.StringVar(value=today_str())
        tk.Label(f_top, text="From:", bg=THEME["bg"], font=FONT_MD).pack(side="left", padx=4)
        tk.Entry(f_top, textvariable=from_date, width=12).pack(side="left")
        tk.Label(f_top, text="To:", bg=THEME["bg"], font=FONT_MD).pack(side="left", padx=4)
        tk.Entry(f_top, textvariable=to_date, width=12).pack(side="left")

        fig = Figure(figsize=(7.5, 5), dpi=100)
        ax = fig.add_subplot(111)
        fig.subplots_adjust(bottom=0.18)

        def load_chart():
            con = db()
            cur = con.cursor()
            cur.execute("""
                        SELECT date, SUM(effective_total) as total_sales
                        FROM sales
                        WHERE date BETWEEN ? AND ?
                        GROUP BY date
                        ORDER BY date
                        """, (from_date.get(), to_date.get()))
            rows = cur.fetchall()
            con.close()

            dates = [r[0] for r in rows]
            totals = [r[1] for r in rows]
            grand_total = sum(totals)

            ax.clear()
            fig.suptitle("")

            if totals:
                ax.plot(dates, totals, marker="o", linestyle="-")
                ax.set_title("Daily Sales Trend", fontsize=12, fontweight="bold")
                ax.set_xlabel("Date")
                ax.set_ylabel("Sales Amount (₹)")
                ax.tick_params(axis="x", rotation=45)

                # Grand total
                fig.suptitle(f"Total Sales: ₹{grand_total:,.2f}",
                             x=0.5, y=0.02, ha="center",
                             fontsize=11, fontweight="bold", color="green")

                # Exported date
                export_date = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ax.text(0.99, 0.99, f"Exported On: {export_date}",
                        ha="right", va="top", transform=ax.transAxes,
                        fontsize=8, color="gray")
            else:
                ax.text(0.5, 0.5, "No sales in selected range",
                        ha="center", va="center", fontsize=12)

            fig.tight_layout()
            canvas.draw()

        canvas = FigureCanvasTkAgg(fig, win)
        canvas.get_tk_widget().pack(fill="both", expand=True)
        canvas.draw()

        # Buttons
        btn_frame = tk.Frame(win, bg=THEME["bg"])
        btn_frame.pack(pady=6)
        tk.Button(btn_frame, text="Apply Filter", command=load_chart).pack(side="left", padx=5)

        def export_chart(fmt):
            save_path = filedialog.asksaveasfilename(
                defaultextension=f".{fmt}",
                filetypes=[(fmt.upper(), f"*.{fmt}")],
                initialfile=f"daily_sales_trend.{fmt}"
            )
            if save_path:
                fig.savefig(save_path, format=fmt)
                messagebox.showinfo("Export", f"Chart exported as {save_path}")

        tk.Button(btn_frame, text="Export PNG", command=lambda: export_chart("png")).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Export JPEG", command=lambda: export_chart("jpeg")).pack(side="left", padx=5)

        # Load first view
        load_chart()
    def export_all_reports(self):
        save_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            initialfile=f"all_reports.pdf"
        )
        if not save_path:
            return

        from matplotlib.backends.backend_pdf import PdfPages
        pdf = PdfPages(save_path)

        def add_chart(fig):
            pdf.savefig(fig)
            plt.close(fig)

        con = db()
        cur = con.cursor()
        today = dt.date.today().isoformat()
        old = "2000-01-01"

        # --- Monthly Sales Trend ---
        cur.execute("""
                    SELECT substr(date,1,7) AS month, SUM(effective_total) as total
                    FROM sales
                    GROUP BY month ORDER BY month
                    """)
        rows = cur.fetchall()
        months = [r[0] for r in rows]
        totals = [r[1] for r in rows]
        fig, ax = plt.subplots()
        ax.plot(months, totals, marker="o")
        ax.set_title("Monthly Sales Trend")
        ax.set_xlabel("Month")
        ax.set_ylabel("Sales Amount (₹)")
        ax.tick_params(axis="x", rotation=45)
        ax.text(0.99, 0.99, f"Exported On: {now_str()}",
                ha="right", va="top", transform=ax.transAxes, fontsize=8, color="gray")
        add_chart(fig)

        # --- Top 5 Products ---
        cur.execute("""
                    SELECT product_name, SUM(quantity) as qty, SUM(effective_total) as sales
                    FROM sales GROUP BY product_id
                    ORDER BY sales DESC LIMIT 5
                    """)
        rows = cur.fetchall()
        names = [r[0] for r in rows]
        qtys = [r[1] for r in rows]
        sales = [r[2] for r in rows]
        fig, ax1 = plt.subplots()
        ax2 = ax1.twinx()
        ax1.bar(names, qtys, color="skyblue", label="Quantity")
        ax2.plot(names, sales, color="orange", marker="o", label="Sales")
        ax1.set_title("Top 5 Products")
        ax1.set_xlabel("Products")
        ax1.set_ylabel("Quantity")
        ax2.set_ylabel("Sales (₹)")
        ax1.tick_params(axis="x", rotation=30)
        ax1.text(0.99, 0.99, f"Exported On: {now_str()}",
                 ha="right", va="top", transform=ax1.transAxes, fontsize=8, color="gray")
        add_chart(fig)

        # --- Supplier vs Supplier ---
        cur.execute("""
                    SELECT supplier_id, SUM(effective_total) as total
                    FROM sales JOIN products p ON sales.product_id=p.product_id
                    GROUP BY supplier_id ORDER BY total DESC
                    """)
        rows = cur.fetchall()
        supp = [r[0] for r in rows]
        totals = [r[1] for r in rows]
        fig, ax = plt.subplots()
        ax.bar(supp, totals, color="green")
        ax.set_title("Supplier vs Supplier Sales")
        ax.set_xlabel("Supplier ID")
        ax.set_ylabel("Total Sales (₹)")
        ax.tick_params(axis="x", rotation=30)
        ax.text(0.99, 0.99, f"Exported On: {now_str()}",
                ha="right", va="top", transform=ax.transAxes, fontsize=8, color="gray")
        add_chart(fig)

        # --- Product Sales Share (Pie) ---
        cur.execute("""
                    SELECT product_name, SUM(effective_total) as total
                    FROM sales WHERE date BETWEEN ? AND ?
                    GROUP BY product_id ORDER BY total DESC
                    """, (old, today))
        rows = cur.fetchall()
        labels = [r[0] for r in rows]
        totals = [r[1] for r in rows]
        fig, ax = plt.subplots()
        if totals:
            ax.pie(totals, labels=labels, autopct="%1.1f%%", startangle=140)
            ax.set_title("Product Sales Share")
            ax.text(0.99, 0.99, f"Exported On: {now_str()}",
                    ha="right", va="top", transform=ax.transAxes, fontsize=8, color="gray")
        add_chart(fig)

        # --- Daily Sales Trend ---
        cur.execute("""
                    SELECT date, SUM(effective_total) as total
                    FROM sales WHERE date BETWEEN ? AND ?
                    GROUP BY date ORDER BY date
                    """, (old, today))
        rows = cur.fetchall()
        dates = [r[0] for r in rows]
        totals = [r[1] for r in rows]
        fig, ax = plt.subplots()
        ax.plot(dates, totals, marker="o")
        ax.set_title("Daily Sales Trend")
        ax.set_xlabel("Date")
        ax.set_ylabel("Sales Amount (₹)")
        ax.tick_params(axis="x", rotation=45)
        ax.text(0.99, 0.99, f"Exported On: {now_str()}",
                ha="right", va="top", transform=ax.transAxes, fontsize=8, color="gray")
        add_chart(fig)

        pdf.close()
        con.close()
        messagebox.showinfo("Export", f"All reports exported:\n{save_path}")
    def show_profit_analysis(self):
        win = tk.Toplevel(self)
        win.title("Profit Analysis Report (Analytical)")
        win.geometry("950x580")

        # --- Date filter & mode selector ---
        f_top = tk.Frame(win, bg=THEME["bg"])
        f_top.pack(fill="x", pady=6)

        from_date = tk.StringVar(value="2000-01-01")
        to_date = tk.StringVar(value=today_str())
        mode = tk.StringVar(value="Daily")

        tk.Label(f_top, text="From:", bg=THEME["bg"], font=FONT_MD).pack(side="left", padx=4)
        tk.Entry(f_top, textvariable=from_date, width=12).pack(side="left")
        tk.Label(f_top, text="To:", bg=THEME["bg"], font=FONT_MD).pack(side="left", padx=4)
        tk.Entry(f_top, textvariable=to_date, width=12).pack(side="left", padx=4)

        tk.Label(f_top, text="Mode:", bg=THEME["bg"], font=FONT_MD).pack(side="left", padx=10)
        ttk.Combobox(f_top, textvariable=mode, values=["Daily", "Monthly"], width=10, state="readonly").pack(side="left")

        # --- Table ---
        cols = ("Period", "Sales (₹)", "Profit (₹)", "% of Total Profit",
                "Growth % vs Prev", "Avg Unit Price (₹)", "Profit Margin %")
        tree = ttk.Treeview(win, columns=cols, show="headings")
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, anchor="center", width=130)
        tree.pack(fill="both", expand=True, pady=10)

        # --- Summary label ---
        summary_lbl = tk.Label(win, text="",
                               font=("Arial", 11, "bold"),
                               fg="blue", bg=THEME["bg"], justify="left")
        summary_lbl.pack(pady=5, anchor="w")

        report_data = []  # store for exports

        def load_report():
            nonlocal report_data
            for row in tree.get_children():
                tree.delete(row)
            report_data = []

            con = db()
            cur = con.cursor()

            if mode.get() == "Daily":
                group_field = "s.date"
            else:
                group_field = "substr(s.date,1,7)"  # YYYY-MM

            query = f"""
                SELECT {group_field} as period,
                       SUM(s.effective_total) AS total_sales,
                       SUM(s.effective_total - (p.unit_price * s.quantity)) AS profit_value,
                       SUM(s.quantity) as total_qty
                FROM sales s
                JOIN products p ON p.product_id = s.product_id
                WHERE s.date BETWEEN ? AND ?
                GROUP BY period
                ORDER BY period
            """
            cur.execute(query, (from_date.get(), to_date.get()))
            rows = cur.fetchall()
            con.close()

            if not rows:
                summary_lbl.config(text="No profit data in this range.")
                return

            total_profit = sum(r[2] for r in rows)
            total_sales = sum(r[1] for r in rows)
            prev_profit = None
            highest_period, lowest_period = None, None
            highest_val, lowest_val = float("-inf"), float("inf")

            for period, sales, profit, qty in rows:
                pct_total = (profit / total_profit * 100) if total_profit else 0
                growth = ((profit - prev_profit) / prev_profit * 100) if prev_profit and prev_profit != 0 else None
                avg_unit_price = (sales / qty) if qty else 0
                profit_margin = (profit / sales * 100) if sales else 0
                prev_profit = profit

                # Track highest/lowest profit
                if profit > highest_val:
                    highest_val, highest_period = profit, period
                if profit < lowest_val:
                    lowest_val, lowest_period = profit, period

                tree.insert("", "end", values=(
                    period,
                    f"₹{sales:,.2f}",
                    f"₹{profit:,.2f}",
                    f"{pct_total:.2f}%",
                    f"{growth:+.2f}%" if growth is not None else "N/A",
                    f"₹{avg_unit_price:,.2f}",
                    f"{profit_margin:.2f}%"
                ))

                report_data.append((period, sales, profit, pct_total, growth, avg_unit_price, profit_margin))

            avg_profit = total_profit / len(rows)
            overall_margin = (total_profit / total_sales * 100) if total_sales else 0

            summary_txt = (
                f"📊 Summary ({mode.get()}):\n"
                f" • Total Sales: ₹{total_sales:,.2f}\n"
                f" • Total Profit: ₹{total_profit:,.2f}\n"
                f" • Overall Profit %: {overall_margin:.2f}%\n"
                f" • Average {mode.get()} Profit: ₹{avg_profit:,.2f}\n"
                f" • Highest Profit: ₹{highest_val:,.2f} in {highest_period}\n"
                f" • Lowest Profit: ₹{lowest_val:,.2f} in {lowest_period}\n"
            )
            summary_lbl.config(text=summary_txt)

        # --- Export Excel ---
        def export_excel():
            from openpyxl import Workbook
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel File", "*.xlsx")],
                initialfile=f"profit_analysis_{mode.get().lower()}.xlsx"
            )
            if save_path:
                wb = Workbook()
                ws = wb.active
                ws.title = f"{mode.get()} Profit Analysis"
                ws.append(cols)
                for row in report_data:
                    ws.append(row)
                ws.append([])
                for line in summary_lbl.cget("text").split("\n"):
                    ws.append([line])
                wb.save(save_path)
                messagebox.showinfo("Export", f"Report saved as {save_path}")

        # --- Export PDF ---
        def export_pdf():
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
            from reportlab.lib.units import mm

            save_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF File", "*.pdf")],
                initialfile=f"profit_analysis_{mode.get().lower()}.pdf"
            )
            if save_path:
                c = canvas.Canvas(save_path, pagesize=A4)
                W, H = A4
                y = H - 30*mm
                c.setFont("Helvetica-Bold", 14)
                c.drawString(50, y, f"Profit Analysis Report ({mode.get()})")
                y -= 20
                c.setFont("Helvetica", 10)
                c.drawString(50, y, f"From {from_date.get()} To {to_date.get()}")
                y -= 20
                c.drawString(50, y, f"Exported On: {dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                y -= 30
                c.setFont("Helvetica-Bold", 8)
                for i, h in enumerate(cols):
                    c.drawString(50 + i*120, y, h)
                y -= 15
                c.setFont("Helvetica", 8)
                for period, sales, profit, pct, growth, avg_unit, margin in report_data:
                    c.drawString(50, y, str(period))
                    c.drawString(170, y, f"₹{sales:,.2f}")
                    c.drawString(290, y, f"₹{profit:,.2f}")
                    c.drawString(410, y, f"{pct:.2f}%")
                    c.drawString(510, y, f"{growth:+.2f}%" if growth is not None else "N/A")
                    c.drawString(610, y, f"₹{avg_unit:,.2f}")
                    c.drawString(710, y, f"{margin:.2f}%")
                    y -= 15
                    if y < 50:
                        c.showPage()
                        y = H - 50
                        c.setFont("Helvetica", 8)
                y -= 20
                c.setFont("Helvetica-Bold", 9)
                for line in summary_lbl.cget("text").split("\n"):
                    c.drawString(50, y, line)
                    y -= 15
                c.save()
                messagebox.showinfo("Export", f"Report saved as {save_path}")

        # --- Buttons ---
        f_btn = tk.Frame(win, bg=THEME["bg"])
        f_btn.pack(pady=5)
        tk.Button(f_btn, text="Apply Filter", command=load_report).pack(side="left", padx=5)
        tk.Button(f_btn, text="Export Excel", command=export_excel).pack(side="left", padx=5)
        tk.Button(f_btn, text="Export PDF", command=export_pdf).pack(side="left", padx=5)

        # First load
        load_report()


# ---------- Run App ----------

if __name__ == "__main__":
    init_db()
    app = InventoryApp()

    app.mainloop()
